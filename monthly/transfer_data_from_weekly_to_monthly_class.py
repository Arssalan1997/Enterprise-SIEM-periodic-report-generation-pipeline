import csv
from monthly import MainSheetOfMonthlyReport
from openpyxl import load_workbook
# from monthly import MainSheetOfWeeklyReportToTransferDataToMonthlyReport   # this import is included later
# in this module, why not applied over here? because of preventing from circular module import issues
from openpyxl.styles import Font, Border, Alignment, PatternFill
from CONSTANTS import FILES_DIRECTORY, MAPPING_FILES_DIRECTORY
import streamlit as st
import os


class TransferDataFromWeeklyToMonthly:

    def __init__(self):

        self.dict_of_mapping_list_of_monthly_to_weekly_sheet_names = dict()    # a dict(list) of tuples with 2 items
        # each pointing to monthly_sheet_name & weekly_sheet_name respectively
        # ###############

        self.number_of_needed_other_excel_files = int()   # in monthly report version of project, it is meant the weekly
        # report files that need to be opened and checked
        self.filenames_of_needed_other_excel_files = list()

        # ###############
        self.list_of_other_workbooks = list()  # refers to the weekly report Excel Spreadsheet files that can be either
        # 2 or 3 files: 2 for 4-week report versions and 3 for 5-week report versions as well

        # ###############

        # will store filenames of weekly report files as keys and the values will be its sheet-names that have
        # issues making it impossible to transfer data to monthly report file
        self.filename_as_key__and__sheet_names_with_issue_as_value = dict()

    def make_mapping_list_of_monthly_to_weekly_sheet_names(self):
        """
        Names of sheets are quite different in some cases between monthly & weekly report corresponding sheet-names.
        That is why this method is made, and the mapping file will be updated at every time that the sheet-names change
        """

        # reference: https://www.programiz.com/python-programming/csv

        # ####### Debugging Lines For The Following 'with open(...)...' #########
        # import os
        # print(os.getcwd())   # == > 'D:\a.bazzaz\SIEM Report Automation With Python Project'
        # #######

        with open(f'{MAPPING_FILES_DIRECTORY}/Monthly Report to Weekly Report Mapping.csv', 'r') as file:
            reader = csv.reader(file)

            # the dict(list) contains tuples with 2 items; first item: sheet-name of monthly report, second
            # item: sheet-name of weekly report. Why '.strip()' applied to 'row[0]' & 'row[1]'? Because there might be
            # cases in which the mapping file gets open manually by users and the
            # record values(monthly&weekly report sheet-names) changed to non-stripped(leading and
            # preceding white-spaces left with the sheet-names of the mapping file)

            self.dict_of_mapping_list_of_monthly_to_weekly_sheet_names =\
                dict([(row[0].strip(), row[1].strip()) for row in reader])

    def determine_number_of_needed_other_excel_files(self):

        if MainSheetOfMonthlyReport.NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS == 4:
            self.number_of_needed_other_excel_files = 2
        elif MainSheetOfMonthlyReport.NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS == 5:
            self.number_of_needed_other_excel_files = 3

        # will later be implemented using GUI
        # print(f"\nPlease enter the '{self.number_of_needed_other_excel_files}' needed weekly excel filenames that are"
        #       f" Located In ./TEST-REPORT Directory Path: ")

        # ####### Debugging Lines For The Following 'with open(...)...' #########
        # import os
        # print(os.getcwd())   # == > 'D:\a.bazzaz\SIEM Report Automation With Python Project'
        # #######

        # # The following snippet receives needed weekly report filenames from user
        # for i in range(self.number_of_needed_other_excel_files):
        #     item = input(f"'{i + 1}'th weekly excel filename = ")
        #     item = f"./TEST-REPORT/{item}"
        #     self.filenames_of_needed_other_excel_files.append(item)

        for weekly_report in st.session_state.weekly_report_files:
            # Construct the full path for the file
            weekly_report_path = os.path.join(FILES_DIRECTORY, weekly_report.name)
            self.filenames_of_needed_other_excel_files.append(weekly_report_path)

            # ------- To make GUI output showing weekly files with their sheet-names that have issues ----------
            basename_of_filename = weekly_report.name    # 'basename_of_filename' will be used as key of the following
            # dictionary
            self.filename_as_key__and__sheet_names_with_issue_as_value[basename_of_filename] = list()  # for each
            # weekly report file ==> the dictionary contains the basename_of_file as the key which is a list that will
            # have the sheets with issues as its items
            # --------------------------------------------------------------------------------------------------

        # test lines of last snippet
        # print(f"\nThe needed weekly excel filenames are as follows : ")
        # for i in self.filenames_of_needed_other_excel_files:
        #     print(i)

    def strip_sheet_names__and_check_if_new_sheet_is_found_in_all_other_excel_spreadsheet_files(self):
        """
        Step 1:

        prevents from issues that can occur in '.copy_data_from_first_other_excel_spreadsheet_to_report_file()'/
        'if monthly_report_sheet.sheet.title == monthly_report_sheet_name' line of code in this module. We do stripping
        on sheet-names of other Excel files before we go for comparing current other file sheet-names with those already
        saved within mapping file[./mappings/Monthly Report to Weekly Report Mapping.csv] because of any probable
        unintended spaces left at starting and ending of other Excel spreadsheet files sheet-names
        # #############################################
        Step 2:

        This step checks the possibility of having new sheets added to other Excel spreadsheet files. if Yes, There will
        most probably be need to add new corresponding sheet and record to the monthly report Excel spreadsheet file and
        also the 'Monthly Report to Weekly Report Mapping.csv' mapping file
        """

        # dictionary including weekly filenames with their sheet-names that are not included within the used mapping
        # file
        # dict_weekly_filename_as_key__and__sheet_names_include_in_file_but_not_in_mapping_file_as_value = dict()
        filename_as_key_and_sheet_names_in_file_but_not_in_mapping_file_as_value = dict()

        for filename in self.filenames_of_needed_other_excel_files:

            # we need the basename of filename(not the fullname)
            basename_of_filename = os.path.basename(filename)

            # the _dict consists of base_filenames as keys ==> made over here and later items will be appended to the
            # list which is the value of the corresponding key
            filename_as_key_and_sheet_names_in_file_but_not_in_mapping_file_as_value[basename_of_filename] = list()

            workbook = load_workbook(filename)   # Why we do not include it within 'try' to handle errors if the
            # filename does not exist? Because we finally implement GUI in which the other Excel files are selected
            # without having to handle FileNotFound Errors.
            sheet_names = workbook.sheetnames    # the job is towards all sheet-names of each Excel spreadsheet file

            # print("#" * 50)
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                sheet.title = sheet_name.strip()   # sheet.title is now changed to 'stripped' format of previous
                # 'sheet_name'

                # ####{{ Step 2 ####

                if sheet.title not in self.dict_of_mapping_list_of_monthly_to_weekly_sheet_names.values():
                    # print("-" * 25)

                    # print(f"\nThe '{filename}' has a sheet with name: '{sheet.title}' that is not found as a "
                    #       f"sheet-name of weekly-report in 'Monthly Report to Weekly Report Mapping.csv' file. "
                    #       f"Please check whether you need to add new corresponding sheet and sheet-name to the "
                    #       f"monthly report Excel spreadsheet file and also the "
                    #       f"'Monthly Report to Weekly Report Mapping.csv' mapping file   "
                    #       f"'transfer_data_from_weekly_to_monthly_class.py': 125")

                    # each key(base filename) has a list of 'sheet.title's as values
                    filename_as_key_and_sheet_names_in_file_but_not_in_mapping_file_as_value[basename_of_filename].append(sheet.title)

                # #### Step 2 }}####
            # print("-" * 50)

            workbook.save(fr"{FILES_DIRECTORY}\{basename_of_filename}")   # added later due to issues of non-stripped
            # sheet-names

            workbook.close()

        from monthly.make_report_info import make_sheet_names_not_found_in_mapping_file

        make_sheet_names_not_found_in_mapping_file(filename_as_key_and_sheet_names_in_file_but_not_in_mapping_file_as_value)

    def provides_other_excel_spreadsheets_to_work_on(self):
        """
        loads weekly report Excel Spreadsheet Files using 'openpyxl' module as we will work on their separate worksheets
        """

        for filename in self.filenames_of_needed_other_excel_files:

            self.list_of_other_workbooks.append(load_workbook(filename))

    @staticmethod
    def delete_needed_number_of_rows_in_monthly_report(monthly_report_sheet, count_of_messages_of_weekly_report,
                                                       number_of_rows_to_delete):
        """
        :param monthly_report_sheet: the tasks are done towards sheets of monthly report spreadsheet file
        :param count_of_messages_of_weekly_report:
        :param number_of_rows_to_delete: the difference between rows in monthly and weekly reports

        ################ %%% ################
        This method takes on 2 related responsibilities:

        1-Doing the deletion
        2-Updating '.list_of_numerical_cells_coordinates' for the very spreadsheet
        """
        # ############## Step 1 ##############

        # detecting the row number from which the rows-removal process must start
        target_cell_coordinate = monthly_report_sheet.list_of_numerical_cells_coordinates[
            count_of_messages_of_weekly_report][0]   # why not the following?

        # ### { ###
        # target_cell_coordinate = monthly_report_sheet.list_of_numerical_cells_coordinates[
        #     count_of_messages_of_weekly_report - 1][0]
        # ### } ###  ==> the target row is the one right after we check same number of rows as messages in weekly report

        row_number_to_start_deletion_from = monthly_report_sheet.sheet[target_cell_coordinate].row
        monthly_report_sheet.sheet.delete_rows(idx=row_number_to_start_deletion_from, amount=number_of_rows_to_delete)

        # ############## Step 2 ##############

        # '.update_all_formulas()' method in 'MainSheetOfMonthlyReport' class works properly as long as
        # '.list_of_numerical_cells_coordinates' remains update before the method is called; now we had some rows
        # deleted in 'monthly_report_sheet', so the 'matrix of coordinates'('.list_of_numerical_cells_coordinates')
        # must change in a way that the coordinates of deleted rows get removed from the matrix, if we do not do this,
        # the formulas will be incorrect.
        [monthly_report_sheet.list_of_numerical_cells_coordinates.pop(-1) for i in range(number_of_rows_to_delete)]

    @staticmethod
    def update_boundary_coordinates_because_row_removal_or_insertion_happened(monthly_report_sheet):
        """
        row removal or insertion happened, and the matrix of numerical cells coordinates
        ('list_of_numerical_cells_coordinates') changed, so we update boundary coordinates that have been changed
        """

        monthly_report_sheet.coordinate_of_down_left_number_boundary = \
            monthly_report_sheet.list_of_numerical_cells_coordinates[-1][0]

        monthly_report_sheet.coordinate_of_down_right_number_boundary = \
            monthly_report_sheet.list_of_numerical_cells_coordinates[-1][-1]

    @staticmethod
    def copy_data_step(monthly_report_sheet, weekly_report_to_be_transferred__sheet):
        """
        We will have 3 steps in this method:

        1-copying message value from weekly report
        2-copying last week numerical value from weekly report
        3-copying this week numerical value from weekly report
        """

        index = 0  # an each-step-incremented variable used to do a loop-like process at the same time with the
        # loop.

        # each 'row' within the loop will contain three first cells in rows of the sheets that contain numerical columns
        # first cell: message value needs to be copied to
        # second cell: last week numerical value needs to be copied to
        # third cell: this week numerical value needs to be copied to
        for row in monthly_report_sheet.sheet.iter_rows(
                min_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].row,
                max_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_left_number_boundary].row,
                min_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].column - 1,
                max_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].column + 1
        ):
            # step 1: message copy
            row[0].value = weekly_report_to_be_transferred__sheet.list_of_messages[index]
            # row[0].value = "test1"
            # step 2: last week num copy
            row[1].value = weekly_report_to_be_transferred__sheet.list_of_last_week_values[index]
            # row[1].value = "test2"
            # step 3: this week num copy
            row[2].value = weekly_report_to_be_transferred__sheet.list_of_this_week_values[index]
            # row[2].value = "test3"

            index += 1  # incremented so that we can have access to the next item within the corresponding lists of
            # 'weekly_report_to_be_transferred__sheet' object

    @staticmethod
    def insert_needed_number_of_rows_in_monthly_report(monthly_report_sheet, number_of_rows_to_insert):
        """
        :param monthly_report_sheet: the tasks are done towards sheets of monthly report spreadsheet file
        :param number_of_rows_to_insert: the difference between rows in weekly and monthly reports

        ################ %%% ################
        This method takes on 3 related responsibilities:

        1-Doing the insertion
        2-Updating '.list_of_numerical_cells_coordinates' for the very spreadsheet
        3-Updating font and style of the new rows that are inserted
        """

        # ############################### Step 1 ###############################

        row_number_to_start_insertion_from = monthly_report_sheet.sheet[
                                                 monthly_report_sheet.coordinate_of_down_left_number_boundary].row + 1

        monthly_report_sheet.sheet.insert_rows(idx=row_number_to_start_insertion_from, amount=number_of_rows_to_insert)

        # ############################### Step 2 ###############################

        temp_list_of_added_row_of_numerical_cells_coordinates = list()   # holds cell coordinates of each row that
        # was inserted in Step 1 of this method

        for row_number_difference in range(1, number_of_rows_to_insert + 1):  # we need 'number_of_rows_to_insert' rows
            # of numerical cells coordinates, the iterator('row_number_difference') in this loop will not be used in
            # its body since we just need to do the following tasks 'number_of_rows_to_insert' number of times

            for coordinate in monthly_report_sheet.list_of_numerical_cells_coordinates[-1]:   # we will use the last
                # list within the matrix of coordinates(which stands for the last rows of coordinates in the matrix) to
                # calculate the following row and column numbers

                # think of the monthly report file to grasp it; the new cell is located right below the last cell on
                # the last row in matrix[last in here means the one already has its coordinate in the matrix]
                # 'row_number_difference': the last row of cell coordinates that is within the matrix of coordinates
                # before this step is our base to add row number in the following line of code

                # Why not the following:
                # 'row_number_of_inserted_cell = monthly_report_sheet.sheet[coordinate].row + row_number_difference'
                # response: the preceding 'for' finds 'monthly_report_sheet.list_of_numerical_cells_coordinates[-1]'
                # dynamically

                row_number_of_inserted_cell = monthly_report_sheet.sheet[coordinate].row + 1
                column_number_of_inserted_cell = monthly_report_sheet.sheet[coordinate].column

                # now that we have the row & column numbers, we can calculate the coordinate of the new cell
                inserted_cell_coordinate = monthly_report_sheet.sheet.cell(
                    row=row_number_of_inserted_cell, column=column_number_of_inserted_cell).coordinate

                temp_list_of_added_row_of_numerical_cells_coordinates.append(inserted_cell_coordinate)

            # we now have 'inserted_cell_coordinate's of each inserted row saved in the temp list, we first append it
            # to the matrix of coordinates and then clear the temp list for next step of the outer list(operations
            # for next inserted row)
            monthly_report_sheet.list_of_numerical_cells_coordinates.append(temp_list_of_added_row_of_numerical_cells_coordinates)
            temp_list_of_added_row_of_numerical_cells_coordinates = []

        # ############################### Step 3 ###############################

        # 1) making the last row before the preceding insertion process as our reference for font and style

        reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows = list()

        # {{
        # ############## Today the message column in monthly report has not non-default font and style, but tomorrow
        # the story might be different. That is why we also include it in reference list and will then try to use for
        # the message columns of newly inserted rows in each sheet of the monthly report file. ##############

        # the reference row is located one row before 'row_number_to_start_insertion_from'
        row_number_of_message_cell_in_last_row_before_newly_inserted_rows = row_number_to_start_insertion_from - 1

        # the message column is located one column before:
        # 'coordinate_of_down_left_number_boundary'['coordinate_of_up_left_number_boundary']
        column_number_of_message_cell_in_last_row_before_newly_inserted_rows = monthly_report_sheet.sheet[
            monthly_report_sheet.coordinate_of_down_left_number_boundary].column - 1

        # cell with the preceding row and column numbers
        message_cell_in_last_row_before_newly_inserted_rows = monthly_report_sheet.sheet.\
            cell(row=row_number_of_message_cell_in_last_row_before_newly_inserted_rows,
                 column=column_number_of_message_cell_in_last_row_before_newly_inserted_rows)

        # the following list, holding the reference row cells, has message cell as its first item
        reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows.append(message_cell_in_last_row_before_newly_inserted_rows)
        # }}

        # ### time to append next items of the reference row list that can be accessed using 'for-loop' ###

        # {{
        # iterates over 'cell_coordinate's in last row before the preceding row-insertions happened

        # why the row with index '[-number_of_rows_to_insert - 1]' is used to iterate over? now that our matrix has been
        # completed in step 2 of this method, we can go back in indexes of the matrix and after
        # 'number_of_rows_to_insert' of backward, we reach to the first-inserted-row, one row before is the target row
        # reference for font and style
        for cell_coordinate in monthly_report_sheet.list_of_numerical_cells_coordinates[-number_of_rows_to_insert - 1]:
            reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows. \
                append(monthly_report_sheet.sheet[cell_coordinate])
        # }}

        # ### time to append 'total' cell to the reference row list ###
        # {{
        # 'total' cell is located after the last item in reference row of cells for font and style
        last_cell_within_reference_row_for_font_and_style = reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows[-1]

        row_number = last_cell_within_reference_row_for_font_and_style.row   # same row as last cell in
        # reference row of cells for font and style
        column_number = last_cell_within_reference_row_for_font_and_style.column + 1  # 'total' cell in
        # row of cells for font and style are located at the right position of last item within
        # 'reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows'.
        cell = monthly_report_sheet.sheet.cell(row=row_number, column=column_number)

        # appending 'total' cell will be the addition before last addition of cells for making
        # reference row of cells for font and style
        reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows.append(cell)
        # }}

        # ### The following steps of appending cells to reference row of cells for font and style is only valid for
        # first three sheets of monthly report file that has 'Aggregation' & 'Total*Aggregation' cells ###

        # {{
        # Note: We consider the three first sheets that have 'Aggregation' & 'Total*Aggregation' columns in
        # monthly report

        # the first three sheets needs the addition of the following items to
        # 'reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows'
        if monthly_report_sheet.__class__.counter_to_objects <= 3:

            # 2 cells are left in this case to be appended: 'Aggregation' & 'Total*Aggregation'; which are located after
            # the last item in 'reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows' to complete
            # the reference row for font and style of first three sheets of monthly report
            last_cell_within_reference_row_for_font_and_style = reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows[-1]

            # 2 cells to append to the row cell ==> 2 iterations for the 'for-loop'
            for i in range(1, 3):
                # using the last cell in 'reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows', we
                # find row & column numbers of 'Aggregation' & 'Total*Aggregation' cells to generate these two cells

                row_number = last_cell_within_reference_row_for_font_and_style.row   # same row as last cell in
                # reference row of cells for font and style
                column_number = last_cell_within_reference_row_for_font_and_style.column + i  # 'Aggregation' &
                # 'Total*Aggregation' cells in row of cells for font and style are located at the right position of
                # last item within 'reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows'.
                cell = monthly_report_sheet.sheet.cell(row=row_number, column=column_number)

                # appending 'Aggregation' & 'Total*Aggregation' cells will be the last addition of cells for making
                # reference row of cells for font and style
                reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows.append(cell)

        # }}                End of making reference row of cells for font and style which is the following list:
        #                   'reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows'

        # ###############################
        # ###############################

        # 2) using the reference row of cells for font and style to change font and style of the newly added rows cells

        # ##### max column of the iteration process through newly added rows becomes different at first three sheets
        # of monthly report since they have also 'Aggregation' & 'Total*Aggregation' columns #####
        # {
        max_column_count_to_apply_font_and_style_to_newly_inserted_rows = \
            monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].column + 1
        # the preceding targets the total column

        # considering first three sheets of monthly report that have 'Aggregation' & 'Total*Aggregation' columns as well
        if monthly_report_sheet.__class__.counter_to_objects <= 3:
            max_column_count_to_apply_font_and_style_to_newly_inserted_rows += 2
        # }

        # Iterates over rows previously inserted.
        # (for three first sheets of monthly report: the 'Aggregation' & 'Total*Aggregation' columns also included)

        # What is the reason behind the value assigned to 'min_col'? x - [1] ==> '1': to include message column as well
        # which is located at left side of first numerical column
        for row_of_cells in monthly_report_sheet.sheet.iter_rows(
                                        min_row=row_number_to_start_insertion_from,
                                        max_row=row_number_to_start_insertion_from + number_of_rows_to_insert - 1,
                                        min_col=column_number_of_message_cell_in_last_row_before_newly_inserted_rows,
                                        max_col=max_column_count_to_apply_font_and_style_to_newly_inserted_rows
                                                                 ):

            # At each row of preceding 'for-loop', we will encounter multiple cells saved within a tuple, we then
            # iterate through the very tuple to work on its cells

            index = 0   # to iterate simultaneously on items of reference row of cells for font and style, increment
            # process is applied inside 'for'

            for target_cell in row_of_cells:
                source_cell = reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows[index]
                index += 1   # to point to next item of the list in next iteration of the loop

                # The following is an exact copy of a code snippet in following:
                # './monthly/monthly_class.py/.update_font_and_style_of_the_new_numerical_column_if_inserted(self)'

                # Copy font
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        vertAlign=source_cell.font.vertAlign,
                        underline=source_cell.font.underline,
                        strike=source_cell.font.strike,
                        color=source_cell.font.color
                    )

                # Copy border
                if source_cell.border:
                    target_cell.border = Border(
                        left=source_cell.border.left,
                        right=source_cell.border.right,
                        top=source_cell.border.top,
                        bottom=source_cell.border.bottom,
                        diagonal=source_cell.border.diagonal,
                        diagonal_direction=source_cell.border.diagonal_direction,
                        outline=source_cell.border.outline,
                        vertical=source_cell.border.vertical,
                        horizontal=source_cell.border.horizontal
                    )

                # Copy alignment
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        text_rotation=source_cell.alignment.text_rotation,
                        wrap_text=source_cell.alignment.wrap_text,
                        shrink_to_fit=source_cell.alignment.shrink_to_fit,
                        indent=source_cell.alignment.indent
                    )

                # Copy fill (color)
                if source_cell.fill:
                    target_cell.fill = PatternFill(
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color,
                        fill_type=source_cell.fill.fill_type
                    )

    def provides_each_sheet_of_other_excel_spreadsheet_to_copy_data_from(self, monthly_report_sheet, index_of_other_workbook):
        """
        provides other weekly report File sheet for data copy after taking several steps

        :param monthly_report_sheet:
        :param index_of_other_workbook: all other(weekly) workbooks are saved in 'self.list_of_other_workbooks', so we
        need to know which one is the target workbook to work on at each call of this method
        :return: 'weekly_report_to_be_transferred__sheet'

        Steps of this method are as follows:

        1-detection of the corresponding 'weekly_report_sheet_name'
        2-'try' to have 'sheet' ready by applying 'sheet = workbook[sheet-name]' openpyxl method
        3-instantiating an object of 'MainSheetOfWeeklyReportToTransferDataToMonthlyReport' class for each sheet of
        weekly report
        4-checking if 'weekly_report_to_be_transferred__sheet' has issue, if yes: output a suitable message and return
        'None'
        5-reaching out to step 5 means that the 'weekly_report_to_be_transferred__sheet' has no issue thus it is okay
        to return it
        """

        #    ##########      STEP 1      ##########

        # why applying '.strip()'? because the following dictionary has already stripped the sheet-names of both monthly
        # and weekly reports.(monthly_report_sheet.sheet.title has not been stripped already)
        if monthly_report_sheet.sheet.title.strip() in self.dict_of_mapping_list_of_monthly_to_weekly_sheet_names.keys():

            monthly_report_sheet_name = monthly_report_sheet.sheet.title
            weekly_report_sheet_name = self.dict_of_mapping_list_of_monthly_to_weekly_sheet_names[monthly_report_sheet_name.strip()]

        #    ##########      STEP 2      ##########

            # if confused about the following: review openpyxl module.
            try:
                sheet = self.list_of_other_workbooks[index_of_other_workbook][weekly_report_sheet_name]

            except KeyError as error:
                # the following error print will finally be included in GUI implementation of the project
                print(error)
                print("-" * 50)
                return None

            except Exception as error:
                # the following error print will finally be included in GUI implementation of the project
                print(error)
                print("-" * 50)
                return None

        #    ##########      STEP 3      ##########

            from monthly import MainSheetOfWeeklyReportToTransferDataToMonthlyReport

            # now 'sheet' is ready for sending as a parameter to the corresponding transfer class, the syntax is
            # same as what can be seen in 'monthly/monthly_class_handler.py/
            # method operations_of_iterations_on_worksheets()'
            weekly_report_to_be_transferred__sheet = MainSheetOfWeeklyReportToTransferDataToMonthlyReport(sheet)

        #    ##########      STEP 4      ##########

            if weekly_report_to_be_transferred__sheet.sheet_has_issue is True:

                # print("-" * 50)
                # the following might be changed when implementing 'sum-if' because the next filenames will be needed
                # then

                # we need the basename of filename(not the fullname), the following key has already been made within
                # 'determine_number_of_needed_other_excel_files()' method, it is a list that we will append item to it.
                basename_of_filename = os.path.basename(self.filenames_of_needed_other_excel_files[index_of_other_workbook])

                self.filename_as_key__and__sheet_names_with_issue_as_value[basename_of_filename].append(weekly_report_to_be_transferred__sheet.sheet.title)

                # print(f"Within weekly report file '{self.filenames_of_needed_other_excel_files[index_of_other_workbook]}', "
                #       f"the sheet '{weekly_report_to_be_transferred__sheet.sheet.title}' has issue which makes it "
                #       f"impossible to transfer its data to monthly report Excel Spreadsheet file, 'transfer_data_from_weekly_to_monthly_class.py': 534")

                return None

        #    ##########      STEP 5      ##########

            return weekly_report_to_be_transferred__sheet

    def copy_data_from_second_weekly_excel_file_to_report_file(self, monthly_report_sheet):
        """
        manages copy operations of second weekly report Excel Spreadsheet File to monthly report Excel Spreadsheet
        File for each sheet of monthly_report. At every call one of monthly report sheets is passed to this method

        :param monthly_report_sheet:  called directly from 'monthly_class_handler.py' ==> 'monthly_report_sheet' is
        delivered first to this method within this class
        :return: None(does the operations towards monthly report file)
        """

        #    ##########      STEP 1      ##########

        weekly_report_to_be_transferred__sheet = self.provides_each_sheet_of_other_excel_spreadsheet_to_copy_data_from\
            (monthly_report_sheet, index_of_other_workbook=0)   # why 'index_of_other_workbook' = 0 ==> as the name of
        # the method suggests, the first other weekly Excel Spreadsheet is target for the copy operations ==> this index
        # will be used in '.provides_each_sheet_of_other_excel_spreadsheet_to_copy_data_from()' method to have access to
        # the first other weekly workbook

        # for the weekly report sheets that have issues, the return value from the last call is 'None' ==> we need to
        # return the 'None' to the first method call (which is from the 'monthly_class_handler.py') to prevent from
        # continuation of execution in this method leading to errors as 'weekly_report_to_be_transferred__sheet' would
        # be 'None'
        if weekly_report_to_be_transferred__sheet is None:
            return None

        #    ##########      STEP 2      ##########

        # ############## Number Of Records Comparison Between Monthly & Weekly Reports ##############
        # There will be 3 different conditions each checked separately as follows:

        # 1)  number of record rows in monthly report > number of records in weekly report.
        # length of '.list_of_numerical_cells_coordinates' stands for the number of records in each report
        if len(monthly_report_sheet.list_of_numerical_cells_coordinates) > \
                len(weekly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates):

            # the following length variable is used to detect from which row in monthly report we need to
            # delete rows
            count_of_messages_of_weekly_report = len(weekly_report_to_be_transferred__sheet.
                                                     list_of_numerical_cells_coordinates)

            # we need to know the difference integer value of last if in order to do the deletion of rows in
            # monthly report spreadsheet file.
            number_of_rows_to_delete = len(monthly_report_sheet.list_of_numerical_cells_coordinates) \
                - len(weekly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates)

            # the following method does the deletion process. the parameters: first: the task is going to
            # affect the 'monthly_report_sheet', second: to detect from which row of 'monthly_report_sheet' the
            # deletion process must start, third: no comment needed
            self.delete_needed_number_of_rows_in_monthly_report(monthly_report_sheet,
                                                                count_of_messages_of_weekly_report,
                                                                number_of_rows_to_delete)
            self.update_boundary_coordinates_because_row_removal_or_insertion_happened(monthly_report_sheet)

            self.copy_data_step(monthly_report_sheet, weekly_report_to_be_transferred__sheet)

        #  #################    #################

        # 2)  number of record rows in monthly report = number of records in weekly report.
        elif len(monthly_report_sheet.list_of_numerical_cells_coordinates) == \
                len(weekly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates):

            self.copy_data_step(monthly_report_sheet, weekly_report_to_be_transferred__sheet)

        #  #################    #################

        # 3)  number of record rows in monthly report < number of records in weekly report.
        elif len(monthly_report_sheet.list_of_numerical_cells_coordinates) < \
                len(weekly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates):

            number_of_rows_to_insert = len(weekly_report_to_be_transferred__sheet.
                                           list_of_numerical_cells_coordinates) - \
                                       len(monthly_report_sheet.list_of_numerical_cells_coordinates)

            self.insert_needed_number_of_rows_in_monthly_report(monthly_report_sheet, number_of_rows_to_insert)

            self.update_boundary_coordinates_because_row_removal_or_insertion_happened(monthly_report_sheet)

            self.copy_data_step(monthly_report_sheet, weekly_report_to_be_transferred__sheet)

    def transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file(self, monthly_report_sheet):
        """
        manages transfer operations of forth weekly report Excel Spreadsheet File to monthly report
        Excel Spreadsheet File for each sheet of monthly_report. At every call one of monthly report sheets is passed
        to this method.

        :param monthly_report_sheet: called directly from 'monthly_class_handler.py' ==> 'monthly_report_sheet' is
        delivered first to this method within this class
        :return: None(does the operations towards monthly report file with no need to return a value)
        """

        #            ##########      STEP 1      ##########

        weekly_report_to_be_transferred__sheet = self.\
            provides_each_sheet_of_other_excel_spreadsheet_to_copy_data_from(monthly_report_sheet, index_of_other_workbook=1)

        # for the weekly report sheets that have issues, the return value from the last call is 'None' ==> we need
        # to return the 'None' to the first method call (which is from the 'monthly_class_handler.py') to prevent
        # from continuation of execution in this method leading to errors as
        # 'weekly_report_to_be_transferred__sheet' would be 'None'
        if weekly_report_to_be_transferred__sheet is None:
            return None

        #            ##########      STEP 2      ##########

        # ########### First Part Of Transferring Data From Forth weekly file to monthly file ###########
        """
        This part of data transfer ==> the same as what we do when we make the report manually: the SUM-IF formula 
        Insertion is what has been created using a non-formula-based approach  
        
        The used method: we will iterate over row records of monthly report file and 
        check whether the message value is in 'weekly_report_to_be_transferred__sheet.list_of_messages'
        
        if yes ==>
        we will have a nested for-loop iterating on 'records of weekly file'[a zip object is just made as the first step
        that represents for 'records of weekly file']to look for the target message; if it is found ==> 
        'last_week_value' and 'this_week_value'[==numerical values of third and forth weekly file sheet] numerical 
         values of the corresponding row record are copied to the right positions in monthly report file(third and forth 
         numerical columns)
        
        if no ==> 
        '0's will be copied to the right positions in monthly report file(third and forth numerical columns)
        """

        # a zip object is made for the simplicity of accessing 'message', 'last_week_value', and 'this_week_value',
        list_zip_of_records_of_weekly_report_sheet = list(zip(
                                                        weekly_report_to_be_transferred__sheet.list_of_messages,
                                                        weekly_report_to_be_transferred__sheet.list_of_last_week_values,
                                                        weekly_report_to_be_transferred__sheet.list_of_this_week_values
                                                              ))

        # the matrix of all numerical columns concatenated by the message column are selected for the loop, the narrow
        # point over here is that we do not need the first and second numerical columns(corresponding for first and
        # second weekly report file numerical values), but for the simplicity of having one-step loop they are also
        # included
        for row in monthly_report_sheet.sheet.iter_rows(
            min_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].row,
            max_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].row,
            min_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].column - 1,
            max_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].column
                                                        ):

            # 'row[0]': the first cell of each row which has the message as its value; if it exists within the target
            # list ==> the process to go for its corresponding 'last_week_value', and 'this_week_value' starts
            if row[0].value in weekly_report_to_be_transferred__sheet.list_of_messages:

                for row_record in list_zip_of_records_of_weekly_report_sheet:

                    if row[0].value == row_record[0]:

                        row[3].value = row_record[1]
                        row[4].value = row_record[2]

                        break   # as soon as the message is found in a 'row_record' of the 'list_zip_...', there will
                        # be no need to continue the inner loop as there will be no 'row_record's afterwards that has
                        # the same message[if the weekly report sheet has non-repeated messages]

            # if not(the message does not exist in target list) ==> '0's filled for the numerical value cells of third
            # and forth weeks
            else:
                row[3].value = 0
                row[4].value = 0

        # ########### Second Part Of Transferring Data From Forth weekly file to monthly file ###########
        """
        This part of data transfer ==> the same as what we do when we make the report manually: the step after  
        SUM-IF formula Insertion: copying data from forth weekly report file(the corresponding sheet) and checking for
        duplicates of messages between monthly report file sheet and weekly report file sheet and then inserting needed
        number of rows to the main report space and then finally transferring corresponding data of the 'new messages'
        to the newly made available space of the report     
        
        The approach towards making the operation happen:
        
        sub-step 1: finding the messages that exist in weekly report file sheet but do not exist in monthly report  
        file sheet
        
        sub-step 2: exception handling which is described later
        
        sub-step 3: finding needed number of rows to insert to monthly report sheet for including messages(with their
        numerical values of third and forth weeks) that exist in weekly report file sheet while do not exist in monthly
        report file sheet
        
        sub-step 4: transferring new messages with their numerical records(the ones not existed in monthly report file  
        sheet before) from forth weekly report file(that also includes third week data) to the monthly report file sheet
        """

        # #################### Sub-Step 1 ####################

        # '.list_of_messages' of 'monthly_report_sheet' is not ready without calling the following method, in other
        # words ==> if '.make_list_of_messages()' is not called, we will have a null '.list_of_messages' for the next
        # list that is going to be made
        monthly_report_sheet.make_list_of_messages()

        # making this list is same as checking for duplicate values between monthly report file sheet and weekly report
        # file sheet, and then selecting out the ones not highlighted which stand for the ones in weekly report sheet
        # but not in monthly report sheet
        messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet = \
            [message for message in weekly_report_to_be_transferred__sheet.list_of_messages if
             message not in monthly_report_sheet.list_of_messages]

        # #################### Sub-Step 2 ####################

        # handling a probable exception: this was seen to happen for some scenarios ==> the target list is cleared if
        # it includes only one string object which starts with the "=SUM("[formula] sub-string.

        # quick note: is there anything problematic with the following exception handling? no, take a look at the
        # following code segment to see that clearing the target list that has only on string-item which has not a
        # valid value ==> not going to make problems

        if len(messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet) == 1 and \
           type(messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet[0]) == str and\
           messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet[0].startswith("=SUM("):

            messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet.clear()

        # #################### Sub-Step 3 ####################

        # items within 'messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet' must be added to the monthly
        # report file sheet main space with their corresponding numerical values of third and forth weeks
        number_of_rows_to_insert = len(messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet)

        # the insertion-related operations makes sense if the following condition is evaluated as 'True'
        if number_of_rows_to_insert >= 1:

            self.insert_needed_number_of_rows_in_monthly_report(monthly_report_sheet, number_of_rows_to_insert)
            self.update_boundary_coordinates_because_row_removal_or_insertion_happened(monthly_report_sheet)

            # #################### Sub-Step 4 ####################

            # the number of rows that are going to be included in following for-loop(the
            # number of iterations) is the same as the number of messages in the following list:
            # 'messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet' ==> so we can do a simultaneous loop
            # with the help of the following helping-index.
            index = 0

            # the matrix of all numerical columns concatenated by the message column are selected for the loop, the rows
            # of the matrix belongs to the newly-inserted cells that are going to be filled with messages and their
            # corresponding numerical values for third and forth week('0's filled as values for cells of first and
            # second week since the message did not exist in the monthly report file sheet before)
            for row in monthly_report_sheet.sheet.iter_rows(
                    min_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_left_number_boundary].row - number_of_rows_to_insert + 1,
                    max_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].row,
                    min_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_left_number_boundary].column - 1,
                    max_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].column
                                                            ):

                # for each newly-inserted row, we have a new message to be filled for the value of message column
                # corresponding cell which will be done after its record is found in the inner for-loop
                message = messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet[index]
                index += 1   # makes it possible to iterate on the items of the list at the same time with the looping
                # on the rows that are newly inserted

                # now that we have one of newly-inserted row with one of the new messages ==> we should look for it
                # within first items of 'list_zip_of_records_of_weekly_report_sheet' and after it is found ==> take the
                # message with numerical values of third and forth weeks and fill the corresponding cell values of
                # monthly report file sheet with the obtained values(first and second weeks have '0's as numerical cell
                # values since the messages are new to the monthly report file sheet)
                for row_record in list_zip_of_records_of_weekly_report_sheet:

                    if row_record[0] == message:

                        row[0].value = message           # filling the cell value with the new 'message'
                        row[1].value = 0                 # filling the cell value with '0'
                        row[2].value = 0                 # filling the cell value with '0'
                        row[3].value = row_record[1]     # filling the cell value with 'last_week_value'
                        row[4].value = row_record[2]     # filling the cell value with 'this_week_value'

                        break   # as soon as the message is found in a 'row_record' of the 'list_zip_...', there will
                        # be no need to continue the inner loop as there will be no 'row_record's afterwards that has
                        # the same message[if the weekly report sheet has non-repeated messages]

    def transfer_data_from_fifth_weekly_excel_file_using_sum_if_formula_to_report_file(self, monthly_report_sheet):
        """
         manages transfer operations of fifth weekly report Excel Spreadsheet File(just numerical values of this
         week column) to monthly report Excel Spreadsheet File for each sheet of monthly_report. At every call one of
         monthly report sheets is passed to this method.

         :param monthly_report_sheet: called directly from 'monthly_class_handler.py' ==> 'monthly_report_sheet' is
         delivered first to this method within this class
         :return: None(does the operations towards monthly report file with no need to return a value)
         """

        #       ##########      STEPS 1 & 2      ##########
        """
        for the monthly report excel files that have 5 weeks ==> the steps until working specifically on the sheets of 
        fifth weekly excel file are the same. that is why we first go for doing the transfer operations on forth weekly 
        excel file and then pursue the control towards working on the fifth weekly excel file
        """

        self.transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file(monthly_report_sheet)

        # #################################################
        # #################################################

        #       ##########      STEP 3      ##########

        weekly_report_to_be_transferred__sheet = self. \
            provides_each_sheet_of_other_excel_spreadsheet_to_copy_data_from(monthly_report_sheet, index_of_other_workbook=2)

        # Detailed Comments: Identical Part in the following method of this class:
        # '.transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()'
        if weekly_report_to_be_transferred__sheet is None:
            return None

        #       ##########      STEP 4      ##########

        # ########### First Part Of Transferring Data From Fifth weekly file to monthly file ###########
        """
        Detailed Comments: Identical Part in the following method of this class:
        '.transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()'
        """

        # a zip object is made for the simplicity of accessing 'message' and 'this_week_value',
        list_zip_of_records_of_weekly_report_sheet = list(zip(
                                                        weekly_report_to_be_transferred__sheet.list_of_messages,
                                                        weekly_report_to_be_transferred__sheet.list_of_this_week_values
                                                              ))

        # Detailed Comments: Identical Part in the following method of this class:
        # '.transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()'
        for row in monthly_report_sheet.sheet.iter_rows(
            min_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].row,
            max_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].row,
            min_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].column - 1,
            max_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].column
                                                        ):

            # 'row[0]': the first cell of each row which has the message as its value; if it exists within the target
            # list ==> the process to go for its corresponding 'this_week_value' starts
            if row[0].value in weekly_report_to_be_transferred__sheet.list_of_messages:

                for row_record in list_zip_of_records_of_weekly_report_sheet:

                    if row[0].value == row_record[0]:

                        row[5].value = row_record[1]

                        break   # as soon as the message is found in a 'row_record' of the 'list_zip_...', there will
                        # be no need to continue the inner loop as there will be no 'row_record's afterwards that has
                        # the same message[if the weekly report sheet has non-repeated messages]

            # if not(the message does not exist in target list) ==> '0' filled for the numerical value cell of fifth
            # week
            else:
                row[5].value = 0

        # ########### Second Part Of Transferring Data From Fifth weekly file to monthly file ###########
        """
        Detailed Comments: Identical Part in the following method of this class:
        '.transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()'
        """

        # #################### Sub-Step 1 ####################

        # '.list_of_messages' of 'monthly_report_sheet' is not ready without calling the following method, in other
        # words ==> if '.make_list_of_messages()' is not called, we will have a null '.list_of_messages' for the next
        # list that is going to be made
        monthly_report_sheet.make_list_of_messages()

        # making this list is same as checking for duplicate values between monthly report file sheet and weekly report
        # file sheet, and then selecting out the ones not highlighted which stand for the ones in weekly report sheet
        # but not in monthly report sheet
        messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet = \
            [message for message in weekly_report_to_be_transferred__sheet.list_of_messages if
             message not in monthly_report_sheet.list_of_messages]

        # #################### Sub-Step 2 ####################

        # Detailed Comments: Identical Part in the following method of this class:
        # '.transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()'

        if len(messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet) == 1 and \
           type(messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet[0]) == str and \
           messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet[0].startswith("=SUM("):

            messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet.clear()

        # #################### Sub-Step 3 ####################

        # items within 'messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet' must be added to the monthly
        # report file sheet main space with their corresponding numerical values of fifth week
        number_of_rows_to_insert = len(messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet)

        # the insertion-related operations makes sense if the following condition is evaluated as 'True'
        if number_of_rows_to_insert >= 1:

            self.insert_needed_number_of_rows_in_monthly_report(monthly_report_sheet, number_of_rows_to_insert)
            self.update_boundary_coordinates_because_row_removal_or_insertion_happened(monthly_report_sheet)

            # #################### Sub-Step 4 ####################

            # Detailed Comments: Identical Part in the following method of this class:
            # '.transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()'
            index = 0

            # Detailed Comments: Identical Part in the following method of this class:
            # '.transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()'
            for row in monthly_report_sheet.sheet.iter_rows(
                min_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_left_number_boundary].row - number_of_rows_to_insert + 1,
                max_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].row,
                min_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_left_number_boundary].column - 1,
                max_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].column
                                                            ):

                # for each newly-inserted row, we have a new message to be filled for the value of message column
                # corresponding cell which will be done after its record is found in the inner for-loop
                message = messages_not_in_monthly_report_sheet_but_in_weekly_report_sheet[index]
                index += 1   # makes it possible to iterate on the items of the list at the same time with the looping
                # on the rows that are newly inserted

                # Detailed Comments: Identical Part in the following method of this class:
                # '.transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()'
                for row_record in list_zip_of_records_of_weekly_report_sheet:

                    if row_record[0] == message:

                        row[0].value = message            # filling the cell value with the new 'message'
                        row[1].value = 0                  # filling the cell value with '0'
                        row[2].value = 0                  # filling the cell value with '0'
                        row[3].value = 0                  # filling the cell value with '0'
                        row[4].value = 0                  # filling the cell value with '0'
                        row[5].value = row_record[1]      # filling the cell value with 'this_week_value'

                        break     # as soon as the message is found in a 'row_record' of the 'list_zip_...', there will
                        # be no need to continue the inner loop as there will be no 'row_record's afterwards that has
                        # the same message[if the weekly report sheet has non-repeated messages]

    def close_weekly_report_excel_spreadsheet_files(self):
        for workbook in self.list_of_other_workbooks:
            workbook.close()

    def send__dictionary_of_filename_as_key__and__sheet_names_with_issue_as_value_to_the_needed_function(self):
        """
        After all iterations of worksheets finished, 'self.filename_as_key__and__sheet_names_with_issue_as_value' is
        completed. We send it to the following function.
        """

        from monthly.make_report_info import make_sheet_names_with_issues_in_weekly_report_files
        make_sheet_names_with_issues_in_weekly_report_files(self.filename_as_key__and__sheet_names_with_issue_as_value)
