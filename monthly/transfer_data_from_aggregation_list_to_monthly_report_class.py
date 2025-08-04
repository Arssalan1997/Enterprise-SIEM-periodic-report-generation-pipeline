from openpyxl import load_workbook
import os
from CONSTANTS import FILES_DIRECTORY_OF_MONTHLY, CURRENT_TIME
import streamlit as st


class TransferDataFromAggregationListToMonthlyReport:

    def __init__(self):
        self.aggregation_list__workbook = None   # will be an object made from using 'load_workbook()' of 'openpyxl'
        # module. why we have made this variable over here? because it is used within several upcoming methods of this
        # class.

        self.info_of_records_with_zero_as_aggregation_value = dict()   # [key: sheet-name, value: message] ==> usage
        # is showing up messages of first 3 sheets of monthly report file to inform the analyzer to check the SIEM for
        # the corresponding message and add to the aggregation list file eventually

    def determine_aggregation_list_excel_file__and__provide_it_to_work_on(self):
        """
        Will later be implemented using GUI
        """

        # aggregation_list__filename = input("\nPlease enter the needed aggregation list filename that is "
        #                                    "located In ./TEST-REPORT Directory Path: ")
        # aggregation_list__filename = f"./TEST-REPORT/{aggregation_list__filename}"

        aggregation_list__filename = os.path.join(fr"{FILES_DIRECTORY_OF_MONTHLY}\{CURRENT_TIME}", st.session_state.aggregation_file.name)

        self.aggregation_list__workbook = load_workbook(aggregation_list__filename)

    def copy_aggregation_values_from_aggregation_list_excel_file_to_monthly_report_file(self, monthly_report_sheet):
        """
        STEP 1: Detecting the sheet-name of first sheet in aggregation list and make it ready to send as an argument to
        'AggregationListSheet' class
        STEP 2: Instantiating an object of 'AggregationListSheet' class and make sure its '.sheet_has_issue' attribute
        is 'False'
        STEP 3: Transferring aggregation values to monthly report file and making the following:
        'self.info_of_records_with_zero_as_aggregation_value'
        """

        #    ##########      STEP 1      ##########

        first_sheet__name = self.aggregation_list__workbook.sheetnames[0]   # the file contains only one sheet which
        # is the case of consideration

        sheet = self.aggregation_list__workbook[first_sheet__name]   # now we have 'sheet' that will be used to access
        # the content of the first sheet of 'Aggregation List.xlsx' file

        #    ##########      STEP 2      ##########

        from monthly import AggregationListSheet

        aggregation_list_sheet__object = AggregationListSheet(sheet)   # the instantiated object will have two
        # attributes that will be used in this class: 1-'.list_of_messages', 2-'.list_of_aggregation_values'

        st.session_state.issue_with_aggregation_list = False   # used to check whether we need to display table of
        # messages with 0_aggregation or not
        if aggregation_list_sheet__object.sheet_has_issue is True:

            st.session_state.issue_with_aggregation_list = True  # a proper message should be displayed about it.

            # print("There is something wrong with the aggregation list,  'transfer_data_from_aggregation_list_to_monthly_report_class.py': 60")

            return None

        #    ##########      STEP 3      ##########

        # initializes a key of the dictionary as a list, it is the title of the sheet in monthly report file
        self.info_of_records_with_zero_as_aggregation_value[monthly_report_sheet.sheet.title] = list()

        # --------------------------------- X

        # a zip object is made for the simplicity of accessing 'message', 'aggregation_value'
        list_zip_of_records_of_aggregation_list_sheet = list(zip(
                                                        aggregation_list_sheet__object.list_of_messages,
                                                        aggregation_list_sheet__object.list_of_aggregation_values
                                                              ))

        # the matrix of all numerical columns concatenated by (the message column and Total & Aggregation columns)
        # are selected for the loop
        for row in monthly_report_sheet.sheet.iter_rows(
                min_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].row,
                max_row=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].row,
                min_col=monthly_report_sheet.sheet[
                            monthly_report_sheet.coordinate_of_up_left_number_boundary].column - 1,
                max_col=monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].column + 2
        ):

            # 'row[0]': the first cell of each row which has the message as its value; if it exists within the target
            # list ==> the process to go for its corresponding 'aggregation_value' starts
            if row[0].value in aggregation_list_sheet__object.list_of_messages:

                for row_record in list_zip_of_records_of_aggregation_list_sheet:

                    if row[0].value == row_record[0]:

                        row[-1].value = row_record[1]

                        break  # as soon as the message is found in a 'row_record' of the 'list_zip_...', there will
                        # be no need to continue the inner loop as there will be no 'row_record's afterwards that has
                        # the same message[if the aggregation list sheet has non-repeated records]

            # if not(the message does not exist in target list) ==> '0's filled for the aggregation value of the
            # corresponding record
            else:
                row[-1].value = 0

                # --------------------------------- X

                # appends record message to the list of the 'monthly_report_sheet.sheet.title' key
                self.info_of_records_with_zero_as_aggregation_value[monthly_report_sheet.sheet.title].append(row[0].value)

    def show_messages_of_monthly_report_sheets_with_zero_value_for_their_aggregation_field(self):

        # There is something wrong with the aggregation file then and no need to go further except closing the file
        if st.session_state.issue_with_aggregation_list is True:
            return None

        # the made dict will be sent to the following imported function for further processing
        from monthly.make_report_info import make_sheet_names_with_messages_that_have_zero_as_aggregation_value_for_sheets_that_have_aggregation_column

        make_sheet_names_with_messages_that_have_zero_as_aggregation_value_for_sheets_that_have_aggregation_column\
            (self.info_of_records_with_zero_as_aggregation_value)

    def close_aggregation_list_excel_spreadsheet_file(self):
        self.aggregation_list__workbook.close()
