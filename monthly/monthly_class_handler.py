"""
Purpose behind making this module a class is to increase readability: we have methods of operations that fall into
four sequential stages, so we defined:

1-'.operations_before_iterations_on_worksheets()'
2-'.operations_of_iterations_on_worksheets()'
3-'.operations_after_iterations_on_worksheets()'
4-'.save_final_file_operation()'

To make a clear distinction between methods belonging to each of the mentioned stages
"""

from monthly import MainSheetOfMonthlyReport
from openpyxl import load_workbook
import xlwings as xw
from CONSTANTS import FILES_DIRECTORY
# import time


class HandlerOfMainSheetOfMonthlyReport:
    """
    Will include 2 steps:

    1-Doing things that are not related to iteration on sheets
    2-Doing iterations on sheets
    """

    def __init__(self, report_type, filename):
        self.report_type = report_type
        self.filename = filename

        self.workbook = None

        self.transfer_data_from_weekly_to_monthly_class__object = None   # the transfer process has multiple steps that
        # can be better managed by a class

        self.select_data_range_of_sheets = list()    # made within '.operations_of_iterations_on_worksheets()' and used
        # in '.operation_of_select_data()' as an argument of 'MainSheetOfMonthlyReport' class.

        self.app = None  # used within '.operation_of_select_data()' to initiate workbook used by xlwings module to
        # implement select-data-step.

        self.wb = None   # used within '.operation_of_select_data()' to initiate workbook used by xlwings module to
        # implement select-data-step.

        self.transfer_data_from_aggregation_list_to_monthly_report__object = None   # will function as an object later
        # in '.operations_before_iterations_on_worksheets()' method; why we have made this variable over here? because
        # it is used within some other upcoming methods of this class.

    def operations_before_iterations_on_worksheets(self):

        # There might be some scenarios in which one month consists of 5 weeks instead of 4, but a quarter has always
        # 3 months['NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS = 3'] and there is no need to take a check
        MainSheetOfMonthlyReport.determine_number_of_last_report_numerical_columns()
        MainSheetOfMonthlyReport.determine_number_of_current_report_numerical_columns()
        MainSheetOfMonthlyReport.receive_header_names_of_numerical_columns_from_user()

        # ###########################################

        from monthly import TransferDataFromAggregationListToMonthlyReport

        # 'TransferDataFromAggregationListToMonthlyReport' class is identical to 'TransferDataFromWeeklyToMonthly'
        # class in terms of the methods implemented inside the class.
        self.transfer_data_from_aggregation_list_to_monthly_report__object = TransferDataFromAggregationListToMonthlyReport()

        self.transfer_data_from_aggregation_list_to_monthly_report__object.determine_aggregation_list_excel_file__and__provide_it_to_work_on()

        # ###########################################

        from monthly import TransferDataFromWeeklyToMonthly

        # 'TransferDataFromWeeklyToMonthly' class has the responsibility of doing all the tasks needed to copy data from
        # the corresponding weekly reports
        self.transfer_data_from_weekly_to_monthly_class__object = TransferDataFromWeeklyToMonthly()

        # now that the intended object is instantiated, it is time to go for the operations implemented within different
        # methods of the 'TransferDataFromWeeklyToMonthly' class
        self.transfer_data_from_weekly_to_monthly_class__object.make_mapping_list_of_monthly_to_weekly_sheet_names()
        self.transfer_data_from_weekly_to_monthly_class__object.determine_number_of_needed_other_excel_files()

        self.transfer_data_from_weekly_to_monthly_class__object. \
            strip_sheet_names__and_check_if_new_sheet_is_found_in_all_other_excel_spreadsheet_files()

        self.transfer_data_from_weekly_to_monthly_class__object.provides_other_excel_spreadsheets_to_work_on()

    def operations_of_iterations_on_worksheets(self):

        self.workbook = load_workbook(self.filename)
        sheet_names = self.workbook.sheetnames  # saved in a list for probable future use

        for sheet_name in sheet_names:
            sheet = self.workbook[sheet_name]
            monthly_report_sheet = MainSheetOfMonthlyReport(sheet)

            # test line
            # print(isinstance(monthly_report_sheet, MainSheetOfMonthlyReport))  # all outputs ==> 'True'

            monthly_report_sheet.remove_values_of_cells_with_numerical_data()
            monthly_report_sheet.delete_message_column_content()
            monthly_report_sheet.insert_or_delete_one_column_if_needed()   # decision on whether to do it or
            # not will be on the part of the called method itself.
            monthly_report_sheet.header_date_fields_update()
            monthly_report_sheet.update_font_and_style_of_the_new_numerical_column_if_inserted()

            # copy-data step from second week report file(which contains copying data from first week as well)
            # note: there are multiple methods that are implemented within the called instance method[encapsulation is
            # met]

            # ## like other instance methods called so far: if the following boolean value is True, it makes no sense to
            # try to do the operations on the very worksheet ##
            if monthly_report_sheet.sheet_has_issue is False:

                self.transfer_data_from_weekly_to_monthly_class__object. \
                    copy_data_from_second_weekly_excel_file_to_report_file(monthly_report_sheet)

                # ############ the method that is called for monthly reports containing 4-week is different from the
                #              monthly reports containing 5-week ############

                if MainSheetOfMonthlyReport.NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS == 4:
                    self.transfer_data_from_weekly_to_monthly_class__object. \
                        transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file(monthly_report_sheet)

                elif MainSheetOfMonthlyReport.NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS == 5:
                    self.transfer_data_from_weekly_to_monthly_class__object. \
                        transfer_data_from_fifth_weekly_excel_file_using_sum_if_formula_to_report_file(monthly_report_sheet)

                """
                # ############## { #############
                -----------------------------------------
                """

                """
                The following is used to make 'self.select_data_range_of_sheets' which is a list containing of tuples
                with 3 items within each tuple: 
                
                1-'monthly_report_sheet.sheet.title'
                2-'coordinate_of_start_cell'
                3-'coordinate_of_end_cell'
                
                'self.select_data_range_of_sheets' ==> will be used later in '.operation_of_select_data()' method of 
                this very class as an argument delivered to '.select_data_for_line_graph()' method of 
                'MainSheetOfMonthlyReport' class.
                """

                # Detecting the row and column number of start cell to go for the 'coordinate_of_start_cell'
                # What is start cell in this context? the cell used in select data part of configuring line graph of
                # the sheet

                row_number_of_start_cell = monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].row - 1
                column_number_of_start_cell = monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_up_left_number_boundary].column - 1

                coordinate_of_start_cell = monthly_report_sheet.sheet.cell(
                                                                        row=row_number_of_start_cell,
                                                                        column=column_number_of_start_cell).coordinate

                # ------------------

                # Detecting the row and column number of end cell to go for the 'coordinate_of_end_cell'
                # What is end cell in this context? the cell used in select data part of configuring line graph of the
                # sheet

                row_number_of_end_cell = monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].row + 1
                column_number_of_end_cell = monthly_report_sheet.sheet[monthly_report_sheet.coordinate_of_down_right_number_boundary].column

                coordinate_of_end_cell = monthly_report_sheet.sheet.cell(
                                                                    row=row_number_of_end_cell,
                                                                    column=column_number_of_end_cell).coordinate

                # consists of a list that has tuples as its items; tuples ==> each sheet 'with no issue' has
                # its title, 'coordinate_of_start_cell' and 'coordinate_of_end_cell' as items of each tuple
                self.select_data_range_of_sheets.append((
                                                        monthly_report_sheet.sheet.title,
                                                        coordinate_of_start_cell,
                                                        coordinate_of_end_cell))

                """
                -----------------------------------------
                # ############## } #############
                """

            if MainSheetOfMonthlyReport.counter_to_objects < 4 and monthly_report_sheet.sheet_has_issue is False:

                self.transfer_data_from_aggregation_list_to_monthly_report__object.\
                 copy_aggregation_values_from_aggregation_list_excel_file_to_monthly_report_file(monthly_report_sheet)

            # ----------------------------------------------------
            monthly_report_sheet.update_all_formulas()

    def operations_after_iterations_on_worksheets(self):

        self.transfer_data_from_weekly_to_monthly_class__object.close_weekly_report_excel_spreadsheet_files()

        self.transfer_data_from_weekly_to_monthly_class__object.\
            send__dictionary_of_filename_as_key__and__sheet_names_with_issue_as_value_to_the_needed_function()

        self.transfer_data_from_aggregation_list_to_monthly_report__object.\
            show_messages_of_monthly_report_sheets_with_zero_value_for_their_aggregation_field()

        self.transfer_data_from_aggregation_list_to_monthly_report__object.close_aggregation_list_excel_spreadsheet_file()

        MainSheetOfMonthlyReport.display_sheets_with_issues()

    def save_file_operation_before_select_data_phase(self):
        """
        We have to close the workbook since there will be a step of automation which is implemented using xlwings
        module of python. The saved file will be reopened for doing xlwings operations in next method of this class
        namely '.operation_of_select_data()'.
        """

        # self.workbook.save(filename="./TEST-REPORT/Report Made of Automation.xlsx")
        self.workbook.save(filename=fr"{FILES_DIRECTORY}\Report Made by Automation System.xlsx")
        self.workbook.close()
        # time.sleep(1)

    def operation_of_select_data(self):
        """
        Workbook of xlwings module is created for the select-data-step of automation, the spreadsheet file that is going
        to be opened here has already been saved and closed within '.save_file_operation_before_select_data_phase()'
        method of this class. After the workbook is ready to work on in this method ==> passed by to
        '.select_data_for_line_graph()' method of 'MainSheetOfMonthlyReport' class for further processing
        """

        # time.sleep(1)
        # Open the Excel file in the background
        self.app = xw.App(visible=False)  # Set visible=False to hide Excel
        # self.wb = self.app.books.open('./TEST-REPORT/Report Made of Automation.xlsx')
        self.wb = self.app.books.open(fr"{FILES_DIRECTORY}\Report Made by Automation System.xlsx")

        # now that 'self.select_data_range_of_sheets' has been made ready to be used as a reference of data ranges of
        # line graph of each sheet, and the xlwings workbook is just initiated ==> these are sent to the following
        # method as arguments.
        MainSheetOfMonthlyReport.select_data_for_line_graph__and__clear_conditional_formatting_rules(self.wb, self.select_data_range_of_sheets)

    def save_final_file_operation_after_select_data_phase(self):
        # Save the workbook
        # self.wb.save("./TEST-REPORT/Report Made of Automation.xlsx")
        self.wb.save(fr"{FILES_DIRECTORY}\Report Made by Automation System.xlsx")

        # Close the workbook
        self.wb.close()

        # Quit the Excel application
        self.app.quit()
