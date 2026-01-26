"""
Contains classes that are used in monthly version of the project:
1-'MainSheetOfMonthlyReport'
2-...
"""

from openpyxl.styles import Font, Border, Alignment, PatternFill
import streamlit as st


class MainSheetOfMonthlyReport:
    # __slots__ = ('sheet',)      # Since there might be a considerable number of objects from the class, using slots
    # can be a helping factor to improve memory usage

    NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS = 4
    NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS = 4

    LIST_HEADER_NAMES_OF_NUMERICAL_COLUMNS = list()  # CONSTANT VARIABLE including dates that will be
    # updated across all worksheets

    counter_to_objects = 0

    sheet_names_with_issues = list()

    def __init__(self, sheet):
        self.__class__.counter_to_objects += 1

        self.sheet = sheet

        self.sheet_has_issue = False  # initial solution to handle exceptions which will improve later; Display the
        # names of such sheets will be required then.

        self.list_of_numerical_cells_coordinates = list()     # A nested list of list, each inner list keeps
        # coordinates of numerical row

        # We will need to know the coordinates of cells starting and ending the first row of numbers, and also the
        # coordinates of cells staring and ending the last row of numbers
        self.coordinate_of_up_left_number_boundary = str()
        self.coordinate_of_up_right_number_boundary = str()
        self.coordinate_of_down_left_number_boundary = str()
        self.coordinate_of_down_right_number_boundary = str()

        self.calculate_all_coordinate_attributes()   # As the name suggests, this instance method will specify the
        # coordinates and sets them to their real values

        self.list_of_messages = list()   # within a certain sheet of the monthly report: message items will all be
        # stored in a list that will later will be compared to the weekly report file messages for data transfer

        # ############################
        from quarterly import MainSheetOfMonthlyReportToTransferDataToQuarterlyReport

        if (self.__class__ == MainSheetOfMonthlyReportToTransferDataToQuarterlyReport) and\
                (self.sheet_has_issue is False):

            self.make_list_of_messages()    # used in Quarterly Report Tool Of The Application ==>
            # 'quarterly\transfer_data_from_monthly_to_quarterly_class.py\copy_data_step()'

            self.list_of_total_values = list()   # for three first sheets: (total * aggregation) values
            self.make_list_of_total_values()     # makes 'list_of_total_values' ready for use within the following:
            # 'quarterly\transfer_data_from_monthly_to_quarterly_class.py\copy_data_step()'

    def calculate_all_coordinate_attributes(self):
        """
        Proper Output of Worksheets With Problems is needed. [the issue with Hoseein's weekly report that made me
        surprised|null cell...]
        """

        from quarterly import MainSheetOfQuarterlyReport

        list_row_record_coordinates = list()   # keep coordinates of each row

        # Iterate through the rows in the sheet
        for row in self.sheet.iter_rows():
            counter_of_numerical_value_in_each_record = 0   # used to prevent from considering the aggregation values,
            # since they are also numerical values

            for cell in row:
                # Check if the row's value is a number
                if isinstance(cell.value, (int, float)):
                    counter_of_numerical_value_in_each_record += 1

                    # This comparison manages to skip the aggregation values that are within the first 3 worksheets
                    # in monthly version of the project
                    if counter_of_numerical_value_in_each_record > \
                            self.__class__.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS:
                        break
                    list_row_record_coordinates.append(cell.coordinate)    # makes the inner list of coordinates
                    # belonging to numerical values in each target row

            if len(list_row_record_coordinates) == self.__class__.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS:
                # in the first row that we have dates and after the rows of numerical values,
                # this list will keep null according to the reset which is applied after this if
                self.list_of_numerical_cells_coordinates.append(list_row_record_coordinates)
                list_row_record_coordinates = []

            elif len(list_row_record_coordinates) == 0:   # for rows with no rows of numbers
                continue

            # the following condition is useful within step of copying data from first other spreadsheet file, "G3"
            # coordinate used to be appended to 'list_row_record_coordinates' in weekly files of having only one record.
            # to better understand we can take a look at the test lines of next else code block
            elif len(list_row_record_coordinates) == 1:
                list_row_record_coordinates = []
                continue

            else:       # for rows with a weired number of numerical values
                self.sheet_has_issue = True

                # before setting this condition, 'self.__class__.sheet_names_with_issues' used to have also sheet-names
                # of sheets with issues of weekly spreadsheet files as well, which is an inheritance matter of fact
                if self.__class__ == MainSheetOfMonthlyReport:
                    self.__class__.sheet_names_with_issues.append(self.sheet.title)

                elif self.__class__ == MainSheetOfQuarterlyReport:
                    self.__class__.sheet_names_with_issues.append(self.sheet.title)

                return None

        # Now that we have the square of coordinates :)
        # (take a look at the Excel spreadsheet file to know why I call it a square), it is time to extract the
        # boundary coordinates out of it.

        try:
            self.coordinate_of_up_left_number_boundary = self.list_of_numerical_cells_coordinates[0][0]
            self.coordinate_of_up_right_number_boundary = self.list_of_numerical_cells_coordinates[0][-1]
            self.coordinate_of_down_left_number_boundary = self.list_of_numerical_cells_coordinates[-1][0]
            self.coordinate_of_down_right_number_boundary = self.list_of_numerical_cells_coordinates[-1][-1]
        except IndexError:
            self.sheet_has_issue = True

            # before setting this condition, 'self.__class__.sheet_names_with_issues' used to have also sheet-names
            # of sheets with issues of weekly spreadsheet files as well, which is an inheritance matter of fact
            if self.__class__ == MainSheetOfMonthlyReport:
                self.__class__.sheet_names_with_issues.append(self.sheet.title)

            elif self.__class__ == MainSheetOfQuarterlyReport:
                self.__class__.sheet_names_with_issues.append(self.sheet.title)

            return None

    @classmethod
    def determine_number_of_last_report_numerical_columns(cls):

        cls.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS = st.session_state.number_of_numerical_columns_of_last_month

    @classmethod
    def determine_number_of_current_report_numerical_columns(cls):

        cls.NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS = st.session_state.number_of_numerical_columns_of_current_month

    @classmethod
    def receive_header_names_of_numerical_columns_from_user(cls):

        cls.LIST_HEADER_NAMES_OF_NUMERICAL_COLUMNS = list(st.session_state.LIST_HEADER_NAMES_OF_NUMERICAL_COLUMNS)

    def remove_values_of_cells_with_numerical_data(self):

        if self.sheet_has_issue is True:
            return None

        # clearing the numerical values of sequent time_based columns

        for row in self.list_of_numerical_cells_coordinates:
            for coordinate in row:
                self.sheet[coordinate].value = ""

        # clearing the aggregation column numbers of the first 3 sheets of the monthly version of the report

        if (self.__class__ == MainSheetOfMonthlyReport) and (self.__class__.counter_to_objects < 4):
            # '<4':The first 3 sheets(Monthly Version) have aggregation columns
            # Take a look at the Excel Spreadsheet to know why the column is added by 2

            aggregation_column_number = self.sheet[self.coordinate_of_up_right_number_boundary].column + 2

            for row in self.sheet.iter_rows(min_row=self.sheet[self.coordinate_of_up_right_number_boundary].row,
                                            max_row=self.sheet[self.coordinate_of_down_right_number_boundary].row,
                                            min_col=aggregation_column_number,
                                            max_col=aggregation_column_number
                                            ):
                row[0].value = ""  # sets the value to null

    def delete_message_column_content(self):

        if self.sheet_has_issue is True:
            return None

        for row in self.sheet.iter_rows(min_row=self.sheet[self.coordinate_of_up_left_number_boundary].row,
                                        max_row=self.sheet[self.coordinate_of_down_left_number_boundary].row,
                                        min_col=self.sheet[self.coordinate_of_up_left_number_boundary].column - 1,
                                        max_col=self.sheet[self.coordinate_of_down_left_number_boundary].column - 1
                                        ):
            row[0].value = ""  # sets the message value to null

    def insert_or_delete_one_column_if_needed(self):
        """
        This instance method contains 2 inner functions:
            1-insert_one_numerical_column()
            2-remove_one_numerical_column()
            3-update_boundary_coordinates_if_insertion_or_removal_happens()

        After the following function definitions we have:
        The needed concept to decide whether none of them or one of them must be called
        """
        def update_boundary_coordinates_if_insertion_or_removal_happens(self_object):
            """
            If either of insertion or removal of a numerical column takes place,
            'self.list_of_numerical_cells_coordinates' changes, so we have to update the corresponding
            boundary coordinates.
            """

            # only these two ones have been changed
            self_object.coordinate_of_up_right_number_boundary = self_object.list_of_numerical_cells_coordinates[0][-1]
            self_object.coordinate_of_down_right_number_boundary = self_object.list_of_numerical_cells_coordinates[-1][-1]

        def insert_one_numerical_column(self_object):
            """
            Although this inner function receives 'self' as its argument, not defined as an instance method, why?

            Response: Decision & Insertion-Removal Operation are handled in one single instance method in this way which
              makes it more readable; the first solution that I implemented was considering the
              'insert_one_numerical_column()' and 'delete_one_numerical_column()' inner functions as instance methods,
              but I changed that to this implementation, one other reason for this change is that we must not be able
              to call them as instance methods(in direct way) without deciding on whether they are needed or not.
              (I mean without calling '.insert_or_delete_one_column_if_needed()' instance method)

            And finally we call 'update_boundary_coordinates_if_insertion_or_removal_happens()' as we need to refer to
            the boundary coordinates later
            """

            # ############ The following 2 lines: column insertion ############
            column_number_to_insert_column_before =\
                self_object.sheet[self_object.coordinate_of_up_right_number_boundary].column + 1

            self_object.sheet.insert_cols(idx=column_number_to_insert_column_before)

            # The following till the end of the function: updating 'self.list_of_numerical_cells_coordinates'
            # for next phases of operation.(when week-report spreadsheets come into play...)

            # Updating 'self.list_of_numerical_cells_coordinates' based on the insertion on 'qualified' worksheets
            for row_of_cell_coordinates in self_object.list_of_numerical_cells_coordinates:

                # specifying row & column number of new added cell
                row_number_of_added_cell = self_object.sheet[row_of_cell_coordinates[-1]].row    # row number remains
                # the same as one cell before
                column_number_of_added_cell = self_object.sheet[row_of_cell_coordinates[-1]].column + 1  # column number
                # increments by 1

                new_cell_coordinate = self_object.sheet. \
                    cell(row=row_number_of_added_cell, column=column_number_of_added_cell).coordinate

                # #################### Appending Step ####################
                row_of_cell_coordinates.append(new_cell_coordinate)
            # test line
            # print(f"'self.list_of_numerical_cells_coordinates' for sheet '{self_object.sheet.title}' equals to : "
            #       f"{self_object.list_of_numerical_cells_coordinates}")
            # print()
            update_boundary_coordinates_if_insertion_or_removal_happens(self_object)
            # ----------------------------------------------------------------------------

        def delete_one_numerical_column(self_object):
            """
            Refer to the documentation of the preceding inner function('.insert_or_delete_one_column_if_needed()')
            """

            # ############ The following 3 lines: column deletion ############
            column_number_to_delete_column = \
                self_object.sheet[self_object.coordinate_of_up_right_number_boundary].column
            self_object.sheet.delete_cols(idx=column_number_to_delete_column)

            # updating 'self.list_of_numerical_cells_coordinates' based on the removal on 'qualified' worksheets
            for row_of_cell_coordinates in self_object.list_of_numerical_cells_coordinates:
                row_of_cell_coordinates.pop()  # all the last coordinate item of each row in matrix needs to be removed

            # test line
            # print(f"'self.list_of_numerical_cells_coordinates' for sheet '{self_object.sheet.title}' equals to : "
            #       f"{self_object.list_of_numerical_cells_coordinates}")
            # print()

            update_boundary_coordinates_if_insertion_or_removal_happens(self_object)
            # ----------------------------------------------------------------------------

        # ###########################################
        if self.sheet_has_issue is True:
            return None

        if self.__class__.NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS \
                == self.__class__.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS:   # both equals to '4' or '5'
            return None   # No change is required in number of numerical columns

        elif self.__class__.NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS == 5 and\
                self.__class__.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS == 4:

            insert_one_numerical_column(self)

        elif self.__class__.NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS == 4 and \
                self.__class__.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS == 5:

            delete_one_numerical_column(self)

    def header_date_fields_update(self):
        """
        Step 2
        """

        if self.sheet_has_issue is True:
            return None

        # one row upper the first number-included row has all dates
        row_of_numerical_columns_header = self.sheet[self.coordinate_of_up_left_number_boundary].row - 1

        # first number-included row has the same column as the first date
        column_of_first_numerical_column_header = self.sheet[self.coordinate_of_up_left_number_boundary].column

        for header_name_of_numerical_column in self.__class__.LIST_HEADER_NAMES_OF_NUMERICAL_COLUMNS:

            self.sheet.cell(row=row_of_numerical_columns_header, column=column_of_first_numerical_column_header)\
                .value = header_name_of_numerical_column

            column_of_first_numerical_column_header += 1

    def update_font_and_style_of_the_new_numerical_column_if_inserted(self):
        """
        After the last instance method is called('.header_date_fields_update()'), we should check whether a new
        numerical column is inserted or not; if the answer would be 'Yes', the new column is not same as the other
        numerical columns in terms of font & style ==> a mapping that sets it same as one cell before is implemented
        over here(each cell in new numerical column is same as its left side one)
        """

        if self.sheet_has_issue is True:
            return None

        # Checking if insertion of new numerical column has taken place or not
        if self.__class__.NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS == 5 and\
                self.__class__.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS == 4:

            # row range: from 'header_date_field' to the row keeping 'SUM' values
            # column range: two last numerical columns
            for row in self.sheet.iter_rows(min_row=self.sheet[self.coordinate_of_up_right_number_boundary].row - 1,
                                            max_row=self.sheet[self.coordinate_of_down_right_number_boundary].row + 1,
                                            min_col=self.sheet[self.coordinate_of_up_right_number_boundary].column - 1,
                                            max_col=self.sheet[self.coordinate_of_up_right_number_boundary].column
                                            ):

                # each 'row' contains a tuple of 2 cells, the first cell is used as the reference to assign
                # font and border from, the second one is the corresponding cell of the newly added numerical column
                source_cell = row[0]
                target_cell = row[1]

                # Copy font
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        vertAlign=source_cell.font.vertAlign,
                        underline=source_cell.font.underline,
                        strike=source_cell.font.strike,
                        color=source_cell.font.color
                    )

                # Copy border
                if source_cell.border:
                    target_cell.border = Border(
                        left=source_cell.border.left,
                        right=source_cell.border.right,
                        top=source_cell.border.top,
                        bottom=source_cell.border.bottom,
                        diagonal=source_cell.border.diagonal,
                        diagonal_direction=source_cell.border.diagonal_direction,
                        outline=source_cell.border.outline,
                        vertical=source_cell.border.vertical,
                        horizontal=source_cell.border.horizontal
                    )

                # Copy alignment
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        text_rotation=source_cell.alignment.text_rotation,
                        wrap_text=source_cell.alignment.wrap_text,
                        shrink_to_fit=source_cell.alignment.shrink_to_fit,
                        indent=source_cell.alignment.indent
                    )

                # Copy fill (color)
                if source_cell.fill:
                    target_cell.fill = PatternFill(
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color,
                        fill_type=source_cell.fill.fill_type
                    )

    def update_all_formulas(self):
        """
        Different Groups Of Formulas To Be Updated:

        1-SUM of numerical columns
        2-SUM of each message numerical records
        3-SUM of SUM_of_numerical_columns(SUM of each SUM in step '2')

        for 3 first worksheets of 'MainSheetOfMonthlyReport' class, we have the following groups as well:
        4-'Total' * 'Aggregation' of each row(message)
        5-SUM of 'Total' * 'Aggregation' values
        """

        if self.sheet_has_issue is True:
            return None

        """
        ########### Implementing '1-SUM of numerical columns'###########
        """

        coordinates_of_sum_cells_of_numerical_columns = list()     # we will first make it and then iterate over it
        # to set the value of each coordinate

        for coordinate in self.list_of_numerical_cells_coordinates[-1]:  # the last list within the matrix contains
            # coordinates of numerical data of last message

            # we save the row and column number of target cell by calling the row and column number of one cell up
            row_number_of_sum_cell = self.sheet[coordinate].row + 1
            column_number_of_sum_cell = self.sheet[coordinate].column

            # now that we have the row and column number of target cell, we can have access to its coordinate as follows
            coordinate_of_sum_cell = self.sheet.\
                cell(row=row_number_of_sum_cell, column=column_number_of_sum_cell).coordinate

            # 'coordinate_of_sum_cell' is ready to be appended to the list
            coordinates_of_sum_cells_of_numerical_columns.append(coordinate_of_sum_cell)

        # ################################ Will be needed within: #######################################
        # ########## '\quarterly\make_needed_formulas_within_first_sheet_of_quarterly_file.py' ##########

        from quarterly import MainSheetOfQuarterlyReport
        from CONSTANTS import SHEET_NAMES__AND__CORRESPONDING_MONTHLY_SUM_COORDINATES

        if self.__class__ == MainSheetOfQuarterlyReport:
            if self.sheet.title.strip() in SHEET_NAMES__AND__CORRESPONDING_MONTHLY_SUM_COORDINATES.keys():
                SHEET_NAMES__AND__CORRESPONDING_MONTHLY_SUM_COORDINATES[self.sheet.title.strip()] = list(coordinates_of_sum_cells_of_numerical_columns)

        # ###############################################################################################

        index = 0   # as the name of the variable suggests, it helps to sort of iterate at the same time on some other
        # iterator other than 'coordinates_of_sum_cells_of_numerical_columns'. Question: so what is that iterator:
        # Response: coordinates in first and last list kept in 'self.list_of_numerical_cells_coordinates'

        for coordinate in coordinates_of_sum_cells_of_numerical_columns:

            # saves needed corresponding coordinates of first&last lists in 'self.list_of_numerical_cells_coordinates'
            first_coordinate_of_numerical_column = self.list_of_numerical_cells_coordinates[0][index]
            last_coordinate_of_numerical_column = self.list_of_numerical_cells_coordinates[-1][index]

            # needed formula inserted as the value of the cell
            self.sheet[coordinate].value =\
                f'=SUM({first_coordinate_of_numerical_column}:{last_coordinate_of_numerical_column})'

            index += 1

        """
        ############ Implementing '2-SUM of each message numerical records' ############
        """

        # we iterate over each list in 'self.list_of_numerical_cells_coordinates', the first and last coordinates are
        # just the first&last items in each inner list, and the target total cell has row and column easily calculated
        # from the last coordinate in each inner list
        for coordinates_of_numerical_row in self.list_of_numerical_cells_coordinates:

            first_coordinate_in_numerical_row = coordinates_of_numerical_row[0]
            last_coordinate_in_numerical_row = coordinates_of_numerical_row[-1]

            # to understand easily why the row number is the same but the column number incremented ==> look at the
            # Excel Spreadsheets Reports
            row_number_of_total_cell = self.sheet[last_coordinate_in_numerical_row].row
            column_number_of_total_cell = self.sheet[last_coordinate_in_numerical_row].column + 1

            # needed formula inserted as the value of the cell
            self.sheet.cell(row=row_number_of_total_cell, column=column_number_of_total_cell).\
                value = f'=SUM({first_coordinate_in_numerical_row}:{last_coordinate_in_numerical_row})'

        """
        ############ Implementing '3-SUM of SUM_of_numerical_columns(SUM of each SUM in step '2')' ############
        """

        # we have made 'coordinates_of_sum_cells_of_numerical_columns' in step 1, the Excel Formula for this part needs
        # the first&last of items within the list
        first_coordinate_of_sum_of_matching_numerical_column = coordinates_of_sum_cells_of_numerical_columns[0]
        last_coordinate_of_sum_of_matching_numerical_column = coordinates_of_sum_cells_of_numerical_columns[-1]

        # to understand the reason behind the calculation, take a look at the reports
        row_number_of_total_of_sums_cell = self.sheet[last_coordinate_of_sum_of_matching_numerical_column].row
        column_number_of_total_of_sums_cell = self.sheet[last_coordinate_of_sum_of_matching_numerical_column].column + 1

        # needed formula inserted as the value of the cell
        self.sheet.cell(row=row_number_of_total_of_sums_cell, column=column_number_of_total_of_sums_cell).\
            value = f'=SUM({first_coordinate_of_sum_of_matching_numerical_column}' \
                    f':{last_coordinate_of_sum_of_matching_numerical_column})'

        """
        # #################################### Implementing steps of '4 & 5' ####################################
        # ####################################                               ####################################
        """
        # ##############################################################################################

        # first condition logic of if: currently 'MainSheetOfQuarterlyReport' class inherits from
        # 'MainSheetOfMonthlyReport', but the quarterly version does not need steps 4&5

        # second condition logic of if: within 'MainSheetOfMonthlyReport' class, only the first 3 worksheets have steps
        # 4&5
        if (self.__class__ == MainSheetOfMonthlyReport) and (self.__class__.counter_to_objects < 4):

            """
            # ############ Implementing '4-'Total' * 'Aggregation' of each row(message)' ############
            """

            # like what we saw in step 2, we need to iterate over 'self.list_of_numerical_cells_coordinates' as all the
            # needed cells are located in corresponding rows of the report:

            # we calculate the coordinates of the 2 needed cells in Excel Multiplication Formula, then we go for the
            # row and column numbers of the target 'total * aggregation' cell
            for coordinates_of_numerical_row in self.list_of_numerical_cells_coordinates:

                # ########## extracting the coordinate of corresponding total cell by: ##########

                # first: calculating the corresponding row and column numbers
                row_number_of_corresponding_total_cell = self.sheet[coordinates_of_numerical_row[-1]].row
                column_number_of_corresponding_total_cell = self.sheet[coordinates_of_numerical_row[-1]].column + 1

                # second: calculating the corresponding coordinate
                coordinate_of_corresponding_total_cell = self.sheet.\
                    cell(row=row_number_of_corresponding_total_cell, column=column_number_of_corresponding_total_cell).\
                    coordinate

                # ##################

                # ########## extracting the coordinate of corresponding aggregation cell by: ##########

                # first: calculating the corresponding row and column numbers
                row_number_of_corresponding_aggregation_cell = row_number_of_corresponding_total_cell
                column_number_of_corresponding_aggregation_cell = column_number_of_corresponding_total_cell + 1

                # second: calculating the corresponding coordinate
                coordinate_of_corresponding_aggregation_cell = self.sheet.\
                    cell(row=row_number_of_corresponding_aggregation_cell,
                         column=column_number_of_corresponding_aggregation_cell).coordinate

                # ##################

                # ########## extracting the row and column numbers of 'total*aggregation' cell ##########

                row_number_of_corresponding_total_multiplied_by_aggregation_cell = \
                    row_number_of_corresponding_aggregation_cell

                column_number_of_corresponding_total_multiplied_by_aggregation_cell = \
                    column_number_of_corresponding_aggregation_cell + 1

                # needed formula inserted as the value of the cell
                self.sheet.cell(row=row_number_of_corresponding_total_multiplied_by_aggregation_cell,
                                column=column_number_of_corresponding_total_multiplied_by_aggregation_cell).\
                    value = f"={coordinate_of_corresponding_total_cell}*{coordinate_of_corresponding_aggregation_cell}"

            """
            # ############ Implementing '5-SUM of 'Total' * 'Aggregation' values' ############
            """

            # ######## extracting the coordinate of first needed cell for the Excel Spreadsheet Formula: ########

            # first: calculating the corresponding row and column numbers
            row_number_of_first_total_multiplied_by_aggregation_cell = self.sheet[self.coordinate_of_up_right_number_boundary].row
            column_number_of_first_total_multiplied_by_aggregation_cell = self.sheet[self.coordinate_of_up_right_number_boundary].column + 3

            # second: calculating the corresponding coordinate
            coordinate_of_first_total_multiplied_by_aggregation_cell = self.sheet.\
                cell(row=row_number_of_first_total_multiplied_by_aggregation_cell,
                     column=column_number_of_first_total_multiplied_by_aggregation_cell).coordinate

            # ################

            # ######## extracting the coordinate of second needed cell for the Excel Spreadsheet Formula: ########

            # first: calculating the corresponding row and column numbers
            row_number_of_last_total_multiplied_by_aggregation_cell = self.sheet[self.coordinate_of_down_right_number_boundary].row
            column_number_of_last_total_multiplied_by_aggregation_cell = self.sheet[self.coordinate_of_down_right_number_boundary].column + 3

            # second: calculating the corresponding coordinate
            coordinate_of_last_total_multiplied_by_aggregation_cell = self.sheet.\
                cell(row=row_number_of_last_total_multiplied_by_aggregation_cell,
                     column=column_number_of_last_total_multiplied_by_aggregation_cell).coordinate

            # ################

            # ########## extracting the row and column numbers of 'SUM of [total*aggregation]' cell ##########

            row_number_of_sum_of_total_multiplied_by_aggregation_cell = \
                row_number_of_last_total_multiplied_by_aggregation_cell + 1

            column_number_of_sum_of_total_multiplied_by_aggregation_cell = \
                column_number_of_last_total_multiplied_by_aggregation_cell

            # needed formula inserted as the value of the cell
            self.sheet.cell(row=row_number_of_sum_of_total_multiplied_by_aggregation_cell,
                            column=column_number_of_sum_of_total_multiplied_by_aggregation_cell).\
                value = f"=SUM({coordinate_of_first_total_multiplied_by_aggregation_cell}" \
                        f":{coordinate_of_last_total_multiplied_by_aggregation_cell})" 

    def make_list_of_messages(self):
        if self.sheet_has_issue is True:
            return None

        for row in self.sheet.iter_rows(min_row=self.sheet[self.coordinate_of_up_left_number_boundary].row,
                                        max_row=self.sheet[self.coordinate_of_down_left_number_boundary].row,
                                        min_col=self.sheet[self.coordinate_of_up_left_number_boundary].column - 1,
                                        max_col=self.sheet[self.coordinate_of_down_left_number_boundary].column - 1
                                        ):
            self.list_of_messages.append(row[0].value)   # every row has only one cell

    @classmethod
    def display_sheets_with_issues(cls):

        # print(f"\n'{len(cls.sheet_names_with_issues)}' sheets remained unchanged since had issues needed to observe.,
        # 'monthly_class.py': 621"
        #       f"The following displays their names: \n")

        st.session_state.list_of_sheet_names_with_issues = list(cls.sheet_names_with_issues)

        # for sheet_name in cls.sheet_names_with_issues:
        #     # print(sheet_name)
        #     st.session_state.list_of_sheet_names_with_issues.append(sheet_name)

        from monthly.make_report_info import make_list_of_sheet_names_with_issues_in_monthly_report_file

        make_list_of_sheet_names_with_issues_in_monthly_report_file(st.session_state.list_of_sheet_names_with_issues)

    @staticmethod
    def select_data_for_line_graph__and__clear_conditional_formatting_rules(wb_xlwings, select_data_range_of_sheets):
        """
        Reference: The method to change chart-data-range of line graph belonging to each sheet has been learned from
        DeepSeek LLM Model.

        we iterate through '.sheets' of 'wb_xlwings'; then for each 'sheet' we do the following:
        0-clear all conditional formatting rules
        1-iterate over 'select_data_range_of_sheets' to find the target sheet
        2-compare title of the 'sheet' within 'wb_xlwings'['sheet.name'] with first item in each record(item) of
        'select_data_range_of_sheets', if it matches we go to step 3.
        3-now that we are meeting the wanted tuple item in 'select_data_range_of_sheets', we go for the chart with its
        name, retrieve needed coordinates from second and third item of the corresponding tuple and then set source data
        for the chart after making the range by using the coordinate variables.
        4-we have the operations completed by then ==> break the inner for-loop since we have found our target tuple
        within 'select_data_range_of_sheets' and done our job

        :param wb_xlwings: as the name suggests, refers to the workbook initiated by xlwings module
        :param select_data_range_of_sheets: for more info about it ==> go to '.operations_of_iterations_on_worksheets()'
        method of 'HandlerOfMainSheetOfMonthlyReport' class(the part before the following code line :)
        'monthly_report_sheet.update_all_formulas()'
        """

        # iterates through each sheet within 'wb_xlwings'
        for sheet in wb_xlwings.sheets:

            # ------------------------  {{  ------------------------

            # Clear all conditional formatting rules
            sheet.api.Cells.FormatConditions.Delete()

            # ------------------------  }}  ------------------------

            # each 'record' in 'select_data_range_of_sheets' is a tuple with 3 items representing:
            # 1-'monthly_report_sheet.sheet.title'
            # 2-'coordinate_of_start_cell' of the sheet and finally 3-'coordinate_of_end_cell' of the sheet
            for record in select_data_range_of_sheets:

                if sheet.name == record[0]:     # 'record[0]' holds 'monthly_report_sheet.sheet.title'

                    chart = sheet.charts['Chart 1']  # 'Chart 1' is the name of the chart, now we make the chart object

                    # second and third items of the record(tuple) are the start and end coordinates respectively
                    coordinate_of_start_cell = record[1]
                    coordinate_of_end_cell = record[2]

                    new_data_range_of_graph = sheet.range(f"{coordinate_of_start_cell}:{coordinate_of_end_cell}")
                    chart.set_source_data(new_data_range_of_graph)

                    break   # as soon as the target record(tuple) is found with the help of the mentioned if-condition,
                    # and then the operations are done towards the corresponding sheet of xlwings ==> we break the inner
                    # for-loop to prevent from the useless continuation of it.

    # ############################

    def make_list_of_total_values(self):
        """
        It will be used if the following condition would be True: 'st.session_state.report_type == "Quarterly"'
        """

        # within the following sheets ==> we use (total*aggregation) column instead of total column
        titles_of_three_first_sheets = ['ArcSightRules', 'ArcSightRulesMY', 'ArcSightRulesBMI']

        if self.sheet.title in titles_of_three_first_sheets:

            for row in self.sheet.iter_rows(min_row=self.sheet[self.coordinate_of_up_left_number_boundary].row,
                                            max_row=self.sheet[self.coordinate_of_down_left_number_boundary].row,
                                            min_col=self.sheet[self.coordinate_of_up_left_number_boundary].column,
                                            max_col=self.sheet[self.coordinate_of_up_right_number_boundary].column + 2
                                            ):
                total = 0

                for i in range(self.__class__.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS):
                    total += row[i].value

                total_multiplied_by_aggregation = total * row[-1].value
                self.list_of_total_values.append(total_multiplied_by_aggregation)

        else:

            for row in self.sheet.iter_rows(min_row=self.sheet[self.coordinate_of_up_left_number_boundary].row,
                                            max_row=self.sheet[self.coordinate_of_down_left_number_boundary].row,
                                            min_col=self.sheet[self.coordinate_of_up_left_number_boundary].column,
                                            max_col=self.sheet[self.coordinate_of_up_right_number_boundary].column
                                            ):
                total = 0

                for i in range(self.__class__.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS):
                    total += row[i].value

                self.list_of_total_values.append(total)
