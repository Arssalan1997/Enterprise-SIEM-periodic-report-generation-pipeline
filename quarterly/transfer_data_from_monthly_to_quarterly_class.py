from monthly import TransferDataFromWeeklyToMonthly
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, PatternFill
from CONSTANTS import FILES_DIRECTORY_OF_QUARTERLY, MAPPING_FILES_DIRECTORY
import streamlit as st
import os
from CONSTANTS import LOGS_DIRECTORY


class TransferDataFromMonthlyToQuarterly(TransferDataFromWeeklyToMonthly):

    def __init__(self):
        super().__init__()
        self.dict_of_mapping_list_of_quarterly_to_monthly_sheet_names = dict()

    def make_mapping_list_of_monthly_to_weekly_sheet_names(self):
        """
        :return: There is no no need for this method in this class.
        """
        return None

    def make_mapping_list_of_quarterly_to_monthly_sheet_names(self):
        """
        Names of sheets are quite different in some cases between quarterly & monthly report corresponding sheet-names.
        That is why this method is made, and the mapping file will be updated at every time that the sheet-names change
        """

        with open(f'{MAPPING_FILES_DIRECTORY}/Quarterly Report to Monthly Report Mapping.csv', 'r') as file:
            reader = csv.reader(file)

            self.dict_of_mapping_list_of_quarterly_to_monthly_sheet_names = \
                dict([(row[0].strip(), row[1].strip()) for row in reader])

    def determine_other_excel_files(self):

        for monthly_report in st.session_state.monthly_report_files:
            # Construct the full path for the file
            monthly_report_path = os.path.join(fr"{FILES_DIRECTORY_OF_QUARTERLY}\{st.session_state.CURRENT_TIME}", monthly_report.name)
            self.filenames_of_needed_other_excel_files.append(monthly_report_path)

            # ------- To make GUI output showing weekly files with their sheet-names that have issues ----------
            basename_of_filename = monthly_report.name  # 'basename_of_filename' will be used as key of the following
            # dictionary
            self.filename_as_key__and__sheet_names_with_issue_as_value[basename_of_filename] = list()  # for each
            # weekly report file ==> the dictionary contains the basename_of_file as the key which is a list that will
            # have the sheets with issues as its items
            # --------------------------------------------------------------------------------------------------

    def strip_sheet_names__and__check_if_new_sheet_is_found_in_all_other_excel_spreadsheet_files(self):

        # dictionary including weekly filenames with their sheet-names that are not included within the used mapping
        # file
        # dict_weekly_filename_as_key__and__sheet_names_include_in_file_but_not_in_mapping_file_as_value = dict()
        filename_as_key_and_sheet_names_in_file_but_not_in_mapping_file_as_value = dict()

        for filename in self.filenames_of_needed_other_excel_files:

            # we need the basename of filename(not the fullname)
            basename_of_filename = os.path.basename(filename)

            # the _dict consists of base_filenames as keys ==> made over here and later items will be appended to the
            # list which is the value of the corresponding key
            filename_as_key_and_sheet_names_in_file_but_not_in_mapping_file_as_value[basename_of_filename] = list()

            workbook = load_workbook(filename)  # Why we do not include it within 'try' to handle errors if the
            # filename does not exist? Because we finally implement GUI in which the other Excel files are selected
            # without having to handle FileNotFound Errors.
            sheet_names = workbook.sheetnames  # the job is towards all sheet-names of each Excel spreadsheet file

            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                sheet.title = sheet_name.strip()  # sheet.title is now changed to 'stripped' format of previous
                # 'sheet_name'

                # ####{{ Step 2 ####

                if sheet.title not in self.dict_of_mapping_list_of_quarterly_to_monthly_sheet_names.values():

                    # each key(base filename) has a list of 'sheet.title's as values
                    filename_as_key_and_sheet_names_in_file_but_not_in_mapping_file_as_value[basename_of_filename].append(sheet.title)

                # #### Step 2 }}####

            # added later due to issues of non-stripped sheet-names
            workbook.save(filename)

            workbook.close()

        from quarterly.make_report_info import make_sheet_names_not_found_in_mapping_file

        make_sheet_names_not_found_in_mapping_file(filename_as_key_and_sheet_names_in_file_but_not_in_mapping_file_as_value)

    def provides_other_excel_spreadsheets_to_work_on(self):
        """
        loads monthly report Excel Spreadsheet Files using 'openpyxl' module as we will work on their separate worksheets
        """

        for filename in self.filenames_of_needed_other_excel_files:
            self.list_of_other_workbooks.append(load_workbook(filename, read_only=True, data_only=True))

    @staticmethod
    def delete_needed_number_of_rows_in_quarterly_report(quarterly_report_sheet, count_of_messages_of_monthly_report,
                                                         number_of_rows_to_delete):
        """
        :param quarterly_report_sheet: the tasks are done towards sheets of quarterly report spreadsheet file
        :param count_of_messages_of_monthly_report:
        :param number_of_rows_to_delete: the difference between rows in quarterly and monthly reports

        ################ %%% ################
        This method takes on 2 related responsibilities:

        1-Doing the deletion
        2-Updating '.list_of_numerical_cells_coordinates' for the very spreadsheet
        """
        # ############## Step 1 ##############

        # detecting the row number from which the rows-removal process must start
        target_cell_coordinate = quarterly_report_sheet.list_of_numerical_cells_coordinates[
            count_of_messages_of_monthly_report][0]  # why not the following?

        # ### { ###
        # target_cell_coordinate = quarterly_report_sheet.list_of_numerical_cells_coordinates[
        #     count_of_messages_of_monthly_report - 1][0]
        # ### } ### ==> the target row is the one right after we check same number of rows as messages in monthly report

        row_number_to_start_deletion_from = quarterly_report_sheet.sheet[target_cell_coordinate].row
        quarterly_report_sheet.sheet.delete_rows(idx=row_number_to_start_deletion_from, amount=number_of_rows_to_delete)

        # ############## Step 2 ##############

        # '.update_all_formulas()' method in 'MainSheetOfMonthlyReport' class works properly as long as
        # '.list_of_numerical_cells_coordinates' remains update before the method is called; now we had some rows
        # deleted in 'quarterly_report_sheet', so the 'matrix of coordinates'('.list_of_numerical_cells_coordinates')
        # must change in a way that the coordinates of deleted rows get removed from the matrix, if we do not do this,
        # the formulas will be incorrect.
        [quarterly_report_sheet.list_of_numerical_cells_coordinates.pop(-1) for i in range(number_of_rows_to_delete)]

    @staticmethod
    def update_boundary_coordinates_because_row_removal_or_insertion_happened(quarterly_report_sheet):
        """
        row removal or insertion happened, and the matrix of numerical cells coordinates
        ('list_of_numerical_cells_coordinates') changed, so we update boundary coordinates that have been changed
        """

        quarterly_report_sheet.coordinate_of_down_left_number_boundary = \
            quarterly_report_sheet.list_of_numerical_cells_coordinates[-1][0]

        quarterly_report_sheet.coordinate_of_down_right_number_boundary = \
            quarterly_report_sheet.list_of_numerical_cells_coordinates[-1][-1]

    @staticmethod
    def copy_data_step(quarterly_report_sheet, monthly_report_to_be_transferred__sheet):
        """
        We will have 2 steps in this method:

        1-copying message value from monthly report
        2-copying total('total * aggregation' for first three sheets) value from monthly report
        """

        index = 0  # an each-step-incremented variable used to do a loop-like process at the same time with the
        # loop.

        # each 'row' within the loop will contain two first cells in rows of the sheets that contain numerical columns
        # first cell: message value needs to be copied to
        # second cell: total('total * aggregation' for first three sheets) value needs to be copied to
        for row in quarterly_report_sheet.sheet.iter_rows(
                min_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_up_left_number_boundary].row,
                max_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_left_number_boundary].row,
                min_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_up_left_number_boundary].column - 1,
                max_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_up_left_number_boundary].column
        ):

            # step 1: message value copy
            row[0].value = monthly_report_to_be_transferred__sheet.list_of_messages[index]

            # step 2: total('total * aggregation' for first three sheets) value copy
            row[1].value = monthly_report_to_be_transferred__sheet.list_of_total_values[index]

            # ------------------------
            index += 1  # incremented so that we can have access to the next item within the corresponding lists of
            # 'monthly_report_to_be_transferred__sheet' object

    @staticmethod
    def insert_needed_number_of_rows_in_quarterly_report(quarterly_report_sheet, number_of_rows_to_insert):
        """
        :param quarterly_report_sheet: the tasks are done towards sheets of quarterly report spreadsheet file
        :param number_of_rows_to_insert: the difference between rows in monthly and quarterly reports

        ################ %%% ################
        This method takes on 3 related responsibilities:

        1-Doing the insertion
        2-Updating '.list_of_numerical_cells_coordinates' for the very spreadsheet
        3-Updating font and style of the new rows that are inserted
        """

        # ############################### Step 1 ###############################

        row_number_to_start_insertion_from = quarterly_report_sheet.sheet[
                                             quarterly_report_sheet.coordinate_of_down_left_number_boundary].row + 1

        quarterly_report_sheet.sheet.insert_rows(idx=row_number_to_start_insertion_from, amount=number_of_rows_to_insert)

        # ############################### Step 2 ###############################

        temp_list_of_added_row_of_numerical_cells_coordinates = list()  # holds cell coordinates of each row that
        # was inserted in Step 1 of this method

        for row_number_difference in range(1, number_of_rows_to_insert + 1):  # we need 'number_of_rows_to_insert' rows
            # of numerical cells coordinates, the iterator('row_number_difference') in this loop will not be used in
            # its body since we just need to do the following tasks 'number_of_rows_to_insert' number of times

            for coordinate in quarterly_report_sheet.list_of_numerical_cells_coordinates[-1]:  # we will use the last
                # list within the matrix of coordinates(which stands for the last rows of coordinates in the matrix) to
                # calculate the following row and column numbers

                row_number_of_inserted_cell = quarterly_report_sheet.sheet[coordinate].row + 1
                column_number_of_inserted_cell = quarterly_report_sheet.sheet[coordinate].column

                # now that we have the row & column numbers, we can calculate the coordinate of the new cell
                inserted_cell_coordinate = quarterly_report_sheet.sheet.cell(
                    row=row_number_of_inserted_cell, column=column_number_of_inserted_cell).coordinate

                temp_list_of_added_row_of_numerical_cells_coordinates.append(inserted_cell_coordinate)

            # we now have 'inserted_cell_coordinate's of each inserted row saved in the temp list, we first append it
            # to the matrix of coordinates and then clear the temp list for next step of the outer list(operations
            # for next inserted row)
            quarterly_report_sheet.list_of_numerical_cells_coordinates.append(
                temp_list_of_added_row_of_numerical_cells_coordinates)
            temp_list_of_added_row_of_numerical_cells_coordinates = []

        # ############################### Step 3 ###############################

        # 1) making the last row before the preceding insertion process as our reference for font and style

        reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows = list()

        # {{
        # the reference row is located one row before 'row_number_to_start_insertion_from'
        row_number_of_message_cell_in_last_row_before_newly_inserted_rows = row_number_to_start_insertion_from - 1

        # the message column is located one column before:
        # 'coordinate_of_down_left_number_boundary'['coordinate_of_up_left_number_boundary']
        column_number_of_message_cell_in_last_row_before_newly_inserted_rows = quarterly_report_sheet.sheet[
            quarterly_report_sheet.coordinate_of_down_left_number_boundary].column - 1

        # cell with the preceding row and column numbers
        message_cell_in_last_row_before_newly_inserted_rows = quarterly_report_sheet.sheet. \
            cell(row=row_number_of_message_cell_in_last_row_before_newly_inserted_rows,
                 column=column_number_of_message_cell_in_last_row_before_newly_inserted_rows)

        # the following list, holding the reference row cells, has message cell as its first item
        reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows.append(message_cell_in_last_row_before_newly_inserted_rows)
        # }}

        # ### time to append next items of the reference row list that can be accessed using 'for-loop' ###

        # {{
        # iterates over 'cell_coordinate's in last row before the preceding row-insertions happened

        for cell_coordinate in quarterly_report_sheet.list_of_numerical_cells_coordinates[-number_of_rows_to_insert - 1]:
            reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows. \
                append(quarterly_report_sheet.sheet[cell_coordinate])
        # }}

        # ### time to append 'total' cell to the reference row list ###
        # {{
        # 'total' cell is located after the last item in reference row of cells for font and style
        last_cell_within_reference_row_for_font_and_style = reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows[-1]

        row_number = last_cell_within_reference_row_for_font_and_style.row  # same row as last cell in
        # reference row of cells for font and style
        column_number = last_cell_within_reference_row_for_font_and_style.column + 1  # 'total' cell in
        # row of cells for font and style are located at the right position of last item within
        # 'reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows'.
        total_cell = quarterly_report_sheet.sheet.cell(row=row_number, column=column_number)

        # appending 'total' cell will be the addition before last addition of cells for making
        # reference row of cells for font and style
        reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows.append(total_cell)
        # }}

        # ###############################
        # ###############################

        # 2) using the reference row of cells for font and style to change font and style of the newly added rows cells

        max_column_count_to_apply_font_and_style_to_newly_inserted_rows = \
            quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].column + 1
        # the preceding targets the total column

        # Iterates over rows previously inserted.

        # What is the reason behind the value assigned to 'min_col'? x - [1] ==> '1': to include message column as well
        # which is located at left side of first numerical column
        for row_of_cells in quarterly_report_sheet.sheet.iter_rows(
                                        min_row=row_number_to_start_insertion_from,
                                        max_row=row_number_to_start_insertion_from + number_of_rows_to_insert - 1,
                                        min_col=column_number_of_message_cell_in_last_row_before_newly_inserted_rows,
                                        max_col=max_column_count_to_apply_font_and_style_to_newly_inserted_rows
                                                                   ):

            index = 0  # to iterate simultaneously on items of reference row of cells for font and style, increment
            # process is applied inside 'for'

            for target_cell in row_of_cells:
                source_cell = reference_cells_of_last_row_for_font_and_style_of_newly_inserted_rows[index]
                index += 1  # to point to next item of the list in next iteration of the loop

                # Copy font
                if source_cell.font:
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        vertAlign=source_cell.font.vertAlign,
                        underline=source_cell.font.underline,
                        strike=source_cell.font.strike,
                        color=source_cell.font.color
                    )

                # Copy border
                if source_cell.border:
                    target_cell.border = Border(
                        left=source_cell.border.left,
                        right=source_cell.border.right,
                        top=source_cell.border.top,
                        bottom=source_cell.border.bottom,
                        diagonal=source_cell.border.diagonal,
                        diagonal_direction=source_cell.border.diagonal_direction,
                        outline=source_cell.border.outline,
                        vertical=source_cell.border.vertical,
                        horizontal=source_cell.border.horizontal
                    )

                # Copy alignment
                if source_cell.alignment:
                    target_cell.alignment = Alignment(
                        horizontal=source_cell.alignment.horizontal,
                        vertical=source_cell.alignment.vertical,
                        text_rotation=source_cell.alignment.text_rotation,
                        wrap_text=source_cell.alignment.wrap_text,
                        shrink_to_fit=source_cell.alignment.shrink_to_fit,
                        indent=source_cell.alignment.indent
                    )

                # Copy fill (color)
                if source_cell.fill:
                    target_cell.fill = PatternFill(
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color,
                        fill_type=source_cell.fill.fill_type
                    )

    def provides_each_sheet_of_other_excel_spreadsheet_to_copy_data_from(self, quarterly_report_sheet, index_of_other_workbook):

        #    ##########      STEP 1      ##########

        # why applying '.strip()'? because the following dictionary has already stripped the sheet-names of both monthly
        # and weekly reports.(quarterly_report_sheet.sheet.title has not been stripped already)
        if quarterly_report_sheet.sheet.title.strip() in self.dict_of_mapping_list_of_quarterly_to_monthly_sheet_names.keys():

            quarterly_report_sheet_name = quarterly_report_sheet.sheet.title
            monthly_report_sheet_name = self.dict_of_mapping_list_of_quarterly_to_monthly_sheet_names[quarterly_report_sheet_name.strip()]

            #    ##########      STEP 2      ##########

            sheet = self.list_of_other_workbooks[index_of_other_workbook][monthly_report_sheet_name]

            #    ##########      STEP 3      ##########

            from quarterly import MainSheetOfMonthlyReportToTransferDataToQuarterlyReport

            try:
                monthly_report_to_be_transferred__sheet = MainSheetOfMonthlyReportToTransferDataToQuarterlyReport\
                    (sheet, st.session_state.monthly_report_files__number_of_numerical_columns[index_of_other_workbook])

            except Exception as error:
                error_logs_filepath = rf'{LOGS_DIRECTORY}\error logs.txt'

                with open(error_logs_filepath, 'a') as f:
                    f.write('\n')
                    f.write(f"{st.session_state.CURRENT_TIME} -- Type Of Report: '{st.session_state.report_type}'"
                            f" -- SheetName Of Other Files: '{monthly_report_sheet_name}' -- error -- '{error}'")

            #    ##########      STEP 4      ##########

                basename_of_filename = os.path.basename(self.filenames_of_needed_other_excel_files[index_of_other_workbook])
                self.filename_as_key__and__sheet_names_with_issue_as_value[basename_of_filename].append(monthly_report_sheet_name)

                return None

            if monthly_report_to_be_transferred__sheet.sheet_has_issue is True:

                basename_of_filename = os.path.basename(self.filenames_of_needed_other_excel_files[index_of_other_workbook])
                self.filename_as_key__and__sheet_names_with_issue_as_value[basename_of_filename].append(monthly_report_sheet_name)

                return None

            #    ##########      STEP 5      ##########

            return monthly_report_to_be_transferred__sheet

    def copy_data_from_first_monthly_excel_file_to_report_file(self, quarterly_report_sheet):

        #    ##########      STEP 1      ##########

        monthly_report_to_be_transferred__sheet = self.provides_each_sheet_of_other_excel_spreadsheet_to_copy_data_from \
            (quarterly_report_sheet, index_of_other_workbook=0)

        if monthly_report_to_be_transferred__sheet is None:
            return None

        #    ##########      STEP 2      ##########

        # ############## Number Of Records Comparison Between Quarterly & Monthly Reports ##############
        # There will be 3 different conditions each checked separately as follows:

        # 1)  number of record rows in quarterly report > number of records in monthly report.
        # length of '.list_of_numerical_cells_coordinates' stands for the number of records in each report
        if len(quarterly_report_sheet.list_of_numerical_cells_coordinates) > \
                len(monthly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates):

            # the following length variable is used to detect from which row in quarterly report we need to
            # delete rows
            count_of_messages_of_monthly_report = len(monthly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates)

            # we need to know the difference integer value of last if in order to do the deletion of rows in
            # quarterly report spreadsheet file.
            number_of_rows_to_delete = len(quarterly_report_sheet.list_of_numerical_cells_coordinates) \
                                       - len(monthly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates)

            # the following method does the deletion process. the parameters: first: the task is going to
            # affect the 'quarterly_report_sheet', second: to detect from which row of 'quarterly_report_sheet' the
            # deletion process must start, third: no comment needed
            self.delete_needed_number_of_rows_in_quarterly_report(quarterly_report_sheet,
                                                                  count_of_messages_of_monthly_report,
                                                                  number_of_rows_to_delete)

            self.update_boundary_coordinates_because_row_removal_or_insertion_happened(quarterly_report_sheet)

            self.copy_data_step(quarterly_report_sheet, monthly_report_to_be_transferred__sheet)

        #  #################    #################

        # 2)  number of record rows in quarterly report = number of records in monthly report.
        elif len(quarterly_report_sheet.list_of_numerical_cells_coordinates) == \
                len(monthly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates):

            self.copy_data_step(quarterly_report_sheet, monthly_report_to_be_transferred__sheet)

        #  #################    #################

        # 3)  number of record rows in quarterly report < number of records in monthly report.
        elif len(quarterly_report_sheet.list_of_numerical_cells_coordinates) < \
                len(monthly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates):

            number_of_rows_to_insert = len(monthly_report_to_be_transferred__sheet.list_of_numerical_cells_coordinates) - \
                                       len(quarterly_report_sheet.list_of_numerical_cells_coordinates)

            self.insert_needed_number_of_rows_in_quarterly_report(quarterly_report_sheet, number_of_rows_to_insert)

            self.update_boundary_coordinates_because_row_removal_or_insertion_happened(quarterly_report_sheet)

            self.copy_data_step(quarterly_report_sheet, monthly_report_to_be_transferred__sheet)

    def copy_data_from_second_monthly_excel_file_to_report_file(self, quarterly_report_sheet):
        """
        manages transfer operations of second monthly report Excel Spreadsheet File to quarterly report
        Excel Spreadsheet File for each sheet of quarterly_report. At every call one of quarterly report sheets is
        passed to this method.

        :param quarterly_report_sheet: called directly from 'quarterly_class_handler.py' ==> 'quarterly_report_sheet' is
        delivered first to this method within this class
        :return: None(does the operations towards quarterly report file with no need to return a value)
        """

        #            ##########      STEP 1      ##########

        monthly_report_to_be_transferred__sheet = self.\
            provides_each_sheet_of_other_excel_spreadsheet_to_copy_data_from(quarterly_report_sheet, index_of_other_workbook=1)

        if monthly_report_to_be_transferred__sheet is None:
            return None

        #            ##########      STEP 2      ##########

        # ########### First Part Of Transferring Data From Second Monthly File To Quarterly File ###########
        """
        Identical To 'transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()' method within
        the parent class
        """

        # a zip object is made for the simplicity of accessing 'message' and 'total_value'.
        list_zip_of_records_of_monthly_report_sheet = list(zip(
            monthly_report_to_be_transferred__sheet.list_of_messages,
            monthly_report_to_be_transferred__sheet.list_of_total_values,
        ))

        for row in quarterly_report_sheet.sheet.iter_rows(
                min_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_up_left_number_boundary].row,
                max_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].row,
                min_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_up_left_number_boundary].column - 1,
                max_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].column
        ):

            # 'row[0]': the first cell of each row which has the message as its value; if it exists within the target
            # list ==> the process to go for its corresponding 'total_value' starts
            if row[0].value in monthly_report_to_be_transferred__sheet.list_of_messages:

                for row_record in list_zip_of_records_of_monthly_report_sheet:

                    if row[0].value == row_record[0]:
                        row[2].value = row_record[1]

                        break  # as soon as the message is found in a 'row_record' of the 'list_zip_...', there will
                        # be no need to continue the inner loop as there will be no 'row_record's afterwards that has
                        # the same message[if the monthly report sheet has non-repeated messages]

            # if not(the message does not exist in target list) ==> '0's filled for the numerical value cells of second
            # month
            else:
                row[2].value = 0

        # ########### Second Part Of Transferring Data From Second Monthly File To Quarterly File ###########

        """
        Identical To 'transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()' method within
        the parent class
        """

        # #################### Sub-Step 1 ####################

        # '.list_of_messages' of 'quarterly_report_sheet' is not ready without calling the following method, in other
        # words ==> if '.make_list_of_messages()' is not called, we will have a null '.list_of_messages' for the next
        # list that is going to be made
        quarterly_report_sheet.make_list_of_messages()

        # making this list is same as checking for duplicate values between quarterly report file sheet and monthly
        # report file sheet, and then selecting out the ones not highlighted which stand for the ones in monthly report
        # sheet but not in quarterly report sheet
        messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet = \
            [message for message in monthly_report_to_be_transferred__sheet.list_of_messages if
             message not in quarterly_report_sheet.list_of_messages]

        # #################### Sub-Step 2 ####################

        # handling a probable exception: this was seen to happen for some scenarios ==> the target list is cleared if
        # it includes only one string object which starts with the "=SUM("[formula] sub-string.

        # quick note: is there anything problematic with the following exception handling? no, take a look at the
        # following code segment to see that clearing the target list that has only on string-item which has not a
        # valid value ==> not going to make problems

        if len(messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet) == 1 and \
                type(messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet[0]) == str and \
                messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet[0].startswith("=SUM("):
            messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet.clear()

        # #################### Sub-Step 3 ####################

        # items within 'messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet' must be added to the
        # quarterly report file sheet main space with their corresponding numerical values of total.
        number_of_rows_to_insert = len(messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet)

        # the insertion-related operations makes sense if the following condition is evaluated as 'True'
        if number_of_rows_to_insert >= 1:

            self.insert_needed_number_of_rows_in_quarterly_report(quarterly_report_sheet, number_of_rows_to_insert)
            self.update_boundary_coordinates_because_row_removal_or_insertion_happened(quarterly_report_sheet)

            # #################### Sub-Step 4 ####################

            # the number of rows that are going to be included in following for-loop(the
            # number of iterations) is the same as the number of messages in the following list:
            # 'messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet' ==> so we can do a simultaneous loop
            # with the help of the following helping-index.
            index = 0

            # the matrix of all numerical columns concatenated by the message column are selected for the loop, the rows
            # of the matrix belongs to the newly-inserted cells that are going to be filled with messages and their
            # corresponding numerical values for total('0's filled as values for cells of first month
            # since the message did not exist in the quarterly report file sheet before)
            for row in quarterly_report_sheet.sheet.iter_rows(
                    min_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_left_number_boundary].row - number_of_rows_to_insert + 1,
                    max_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].row,
                    min_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_left_number_boundary].column - 1,
                    max_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].column
            ):

                # for each newly-inserted row, we have a new message to be filled for the value of message column
                # corresponding cell which will be done after its record is found in the inner for-loop
                message = messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet[index]
                index += 1  # makes it possible to iterate on the items of the list at the same time with the looping
                # on the rows that are newly inserted

                # now that we have one of newly-inserted row with one of the new messages ==> we should look for it
                # within first items of 'list_zip_of_records_of_monthly_report_sheet' and after it is found ==> take the
                # message with numerical values of total and fill the corresponding cell values of
                # quarterly report file sheet with the obtained values(first month has '0's as numerical
                # cell values since the messages are new to the quarterly report file sheet)
                for row_record in list_zip_of_records_of_monthly_report_sheet:

                    if row_record[0] == message:
                        row[0].value = message  # filling the cell value with the new 'message'
                        row[1].value = 0  # filling the cell value with '0'
                        row[2].value = row_record[1]  # filling the cell value with 'total_value'

                        break  # as soon as the message is found in a 'row_record' of the 'list_zip_...', there will
                        # be no need to continue the inner loop as there will be no 'row_record's afterwards that has
                        # the same message[if the monthly report sheet has non-repeated messages]

    def copy_data_from_third_monthly_excel_file_to_report_file(self, quarterly_report_sheet):
        """
        manages transfer operations of third monthly report Excel Spreadsheet File to quarterly report
        Excel Spreadsheet File for each sheet of quarterly_report. At every call one of quarterly report sheets is
        passed to this method.

        :param quarterly_report_sheet: called directly from 'quarterly_class_handler.py' ==> 'quarterly_report_sheet' is
        delivered first to this method within this class
        :return: None(does the operations towards quarterly report file with no need to return a value)
        """

        #            ##########      STEP 1      ##########

        monthly_report_to_be_transferred__sheet = self.\
            provides_each_sheet_of_other_excel_spreadsheet_to_copy_data_from(quarterly_report_sheet, index_of_other_workbook=2)

        if monthly_report_to_be_transferred__sheet is None:
            return None

        #            ##########      STEP 2      ##########

        # ########### First Part Of Transferring Data From Third Monthly File To Quarterly File ###########
        """
        Identical To 'transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()' method within
        the parent class
        """

        # a zip object is made for the simplicity of accessing 'message' and 'total_value'.
        list_zip_of_records_of_monthly_report_sheet = list(zip(
            monthly_report_to_be_transferred__sheet.list_of_messages,
            monthly_report_to_be_transferred__sheet.list_of_total_values,
        ))

        for row in quarterly_report_sheet.sheet.iter_rows(
                min_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_up_left_number_boundary].row,
                max_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].row,
                min_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_up_left_number_boundary].column - 1,
                max_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].column
        ):

            # 'row[0]': the first cell of each row which has the message as its value; if it exists within the target
            # list ==> the process to go for its corresponding 'total_value' starts
            if row[0].value in monthly_report_to_be_transferred__sheet.list_of_messages:

                for row_record in list_zip_of_records_of_monthly_report_sheet:

                    if row[0].value == row_record[0]:
                        row[3].value = row_record[1]

                        break  # as soon as the message is found in a 'row_record' of the 'list_zip_...', there will
                        # be no need to continue the inner loop as there will be no 'row_record's afterwards that has
                        # the same message[if the monthly report sheet has non-repeated messages]

            # if not(the message does not exist in target list) ==> '0's filled for the numerical value cells of third
            # month
            else:
                row[3].value = 0

        # ########### Second Part Of Transferring Data From Third Monthly File To Quarterly File ###########

        """
        Identical To 'transfer_data_from_forth_weekly_excel_file_using_sum_if_formula_to_report_file()' method within
        the parent class
        """

        # #################### Sub-Step 1 ####################

        # '.list_of_messages' of 'quarterly_report_sheet' is not ready without calling the following method, in other
        # words ==> if '.make_list_of_messages()' is not called, we will have a null '.list_of_messages' for the next
        # list that is going to be made
        quarterly_report_sheet.make_list_of_messages()

        # making this list is same as checking for duplicate values between quarterly report file sheet and monthly
        # report file sheet, and then selecting out the ones not highlighted which stand for the ones in monthly report
        # sheet but not in quarterly report sheet
        messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet = \
            [message for message in monthly_report_to_be_transferred__sheet.list_of_messages if
             message not in quarterly_report_sheet.list_of_messages]

        # #################### Sub-Step 2 ####################

        # handling a probable exception: this was seen to happen for some scenarios ==> the target list is cleared if
        # it includes only one string object which starts with the "=SUM("[formula] sub-string.

        # quick note: is there anything problematic with the following exception handling? no, take a look at the
        # following code segment to see that clearing the target list that has only on string-item which has not a
        # valid value ==> not going to make problems

        if len(messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet) == 1 and \
                type(messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet[0]) == str and \
                messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet[0].startswith("=SUM("):
            messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet.clear()

        # #################### Sub-Step 3 ####################

        # items within 'messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet' must be added to the
        # quarterly report file sheet main space with their corresponding numerical values of total.
        number_of_rows_to_insert = len(messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet)

        # the insertion-related operations makes sense if the following condition is evaluated as 'True'
        if number_of_rows_to_insert >= 1:

            self.insert_needed_number_of_rows_in_quarterly_report(quarterly_report_sheet, number_of_rows_to_insert)
            self.update_boundary_coordinates_because_row_removal_or_insertion_happened(quarterly_report_sheet)

            # #################### Sub-Step 4 ####################

            # the number of rows that are going to be included in following for-loop(the
            # number of iterations) is the same as the number of messages in the following list:
            # 'messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet' ==> so we can do a simultaneous loop
            # with the help of the following helping-index.
            index = 0

            # the matrix of all numerical columns concatenated by the message column are selected for the loop, the rows
            # of the matrix belongs to the newly-inserted cells that are going to be filled with messages and their
            # corresponding numerical values for total('0's filled as values for cells of first and second month
            # since the message did not exist in the quarterly report file sheet before)
            for row in quarterly_report_sheet.sheet.iter_rows(
                    min_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_left_number_boundary].row - number_of_rows_to_insert + 1,
                    max_row=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].row,
                    min_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_left_number_boundary].column - 1,
                    max_col=quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].column
            ):

                # for each newly-inserted row, we have a new message to be filled for the value of message column
                # corresponding cell which will be done after its record is found in the inner for-loop
                message = messages_not_in_quarterly_report_sheet_but_in_monthly_report_sheet[index]
                index += 1  # makes it possible to iterate on the items of the list at the same time with the looping
                # on the rows that are newly inserted

                # now that we have one of newly-inserted row with one of the new messages ==> we should look for it
                # within first items of 'list_zip_of_records_of_monthly_report_sheet' and after it is found ==> take the
                # message with numerical values of total and fill the corresponding cell values of
                # quarterly report file sheet with the obtained values(first and second months have '0's as numerical
                # cell values since the messages are new to the quarterly report file sheet)
                for row_record in list_zip_of_records_of_monthly_report_sheet:

                    if row_record[0] == message:
                        row[0].value = message  # filling the cell value with the new 'message'
                        row[1].value = 0  # filling the cell value with '0'
                        row[2].value = 0  # filling the cell value with '0'
                        row[3].value = row_record[1]  # filling the cell value with 'total_value'

                        break  # as soon as the message is found in a 'row_record' of the 'list_zip_...', there will
                        # be no need to continue the inner loop as there will be no 'row_record's afterwards that has
                        # the same message[if the monthly report sheet has non-repeated messages]

    def close_monthly_report_excel_spreadsheet_files(self):
        for workbook in self.list_of_other_workbooks:
            workbook.close()

    def send__dictionary_of_filename_as_key__and__sheet_names_with_issue_as_value_to_the_needed_function(self):

        from quarterly.make_report_info import make_sheet_names_with_issues_in_monthly_report_files
        make_sheet_names_with_issues_in_monthly_report_files(self.filename_as_key__and__sheet_names_with_issue_as_value)
