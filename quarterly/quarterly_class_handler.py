from monthly import HandlerOfMainSheetOfMonthlyReport
from quarterly import MainSheetOfQuarterlyReport
from openpyxl import load_workbook
# import xlwings as xw
from CONSTANTS import FILES_DIRECTORY_OF_QUARTERLY
# import streamlit as st


class HandlerOfMainSheetOfQuarterlyReport(HandlerOfMainSheetOfMonthlyReport):

    def __init__(self, report_type, filename):
        super().__init__(report_type, filename)
        self.transfer_data_from_monthly_to_quarterly_class__object = None

    def operations_before_iterations_on_worksheets(self):
        MainSheetOfQuarterlyReport.receive_report_persian_name_to_include_in_first_sheet_from_user()
        MainSheetOfQuarterlyReport.receive_header_names_of_numerical_columns_from_user()

        # ###########################################
        # 'TransferDataFromAggregationListToMonthlyReport' is not used in quarterly version of the project
        # ###########################################

        # ########################################### {{ ###########################################
        """
        What To Include:
        
        'TransferDataFromMonthlyToQuarterly' class must be developed
        """
        # ########################################### }} ###########################################

    def operations_of_iterations_on_worksheets(self):
        """
        Differences towards the base class corresponding method:

        1-calling '.receive_filename_and_save_needed_values_on_first_sheet()' method
        2-calling 'update_first_sheet_info()' method on the first worksheet
        """

        # The following step belongs to the '.operations_before_iterations_on_worksheets()' method. The reason behind
        # the inclusion over here: 1-needing 'self.filename' to send to the following method; 2-simplicity of
        # inheritance

        # 'MainSheetOfMonthlyReportHandler' class does not have the following
        MainSheetOfQuarterlyReport.extract_needed_values_from_first_sheet(self.filename)

        self.workbook = load_workbook(self.filename)
        sheet_names = self.workbook.sheetnames  # saved in a list for probable future use

        counter = 0
        for sheet_name in sheet_names:
            counter += 1
            sheet = self.workbook[sheet_name]
            quarterly_report_sheet = MainSheetOfQuarterlyReport(sheet)
            # test line
            # print(isinstance(main_sheet, MainSheetOfQuarterlyReport))  # all outputs ==> 'True'
            quarterly_report_sheet.remove_values_of_cells_with_numerical_data()
            quarterly_report_sheet.delete_message_column_content()

            if counter == 1:
                # 'MainSheetOfMonthlyReportHandler' class does not have the following
                quarterly_report_sheet.update_first_sheet_info()

            quarterly_report_sheet.header_date_fields_update()

            if (quarterly_report_sheet.sheet_has_issue is False) and (counter >= 3):

                # ########################################### {{ ###########################################
                """
                What To Include:

                Data Transfer From Monthly Report Files To Quarterly Report Files Operations
                """
                # ########################################### }} ###########################################

                """
                The following is used to make 'self.select_data_range_of_sheets' which is a list containing tuples
                with 3 items within each tuple.
                """
                row_number_of_start_cell = quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_up_left_number_boundary].row - 1
                column_number_of_start_cell = quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_up_left_number_boundary].column - 1
                coordinate_of_start_cell = quarterly_report_sheet.sheet.cell(
                                                                             row=row_number_of_start_cell,
                                                                             column=column_number_of_start_cell).coordinate

                row_number_of_end_cell = quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].row + 1
                column_number_of_end_cell = quarterly_report_sheet.sheet[quarterly_report_sheet.coordinate_of_down_right_number_boundary].column
                coordinate_of_end_cell = quarterly_report_sheet.sheet.cell(
                                                                           row=row_number_of_end_cell,
                                                                           column=column_number_of_end_cell).coordinate

                self.select_data_range_of_sheets.append((
                    quarterly_report_sheet.sheet.title,
                    coordinate_of_start_cell,
                    coordinate_of_end_cell))

            # ###########################################
            # 'TransferDataFromAggregationListToMonthlyReport' is not used in quarterly version of the project
            # ###########################################

            quarterly_report_sheet.update_all_formulas()

    def operations_after_iterations_on_worksheets(self):

        # ########################################### {{ ###########################################
        """
        What To Include:

        Data Transfer From Monthly Report Files To Quarterly Report Files Operations ==>
        1-closing monthly report files
        2-preparing related output dictionary
        """
        # ########################################### }} ###########################################

        # ###########################################
        # 'TransferDataFromAggregationListToMonthlyReport' is not used in quarterly version of the project
        # ###########################################

        MainSheetOfQuarterlyReport.display_sheets_with_issues()

    def save_file_operation_before_select_data_phase(self, files_directory_of_report=FILES_DIRECTORY_OF_QUARTERLY):
        super().save_file_operation_before_select_data_phase(files_directory_of_report)

    def operation_of_select_data(self, files_directory_of_report=FILES_DIRECTORY_OF_QUARTERLY):
        super().operation_of_select_data(files_directory_of_report)

    def save_final_file_operation_after_select_data_phase(self, files_directory_of_report=FILES_DIRECTORY_OF_QUARTERLY):
        super().save_final_file_operation_after_select_data_phase(files_directory_of_report)
