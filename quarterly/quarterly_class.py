"""
Contains classes that are used in quarterly version of the project:
1-'MainSheetOfQuarterlyReport' that inherits from 'MainSheetOfMonthlyReport' class
2-...
"""

from monthly import MainSheetOfMonthlyReport
from openpyxl import load_workbook


class MainSheetOfQuarterlyReport(MainSheetOfMonthlyReport):

    NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS = 3

    # This class must not have 'NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS' as defined in 'MainSheetOfMonthlyReport'
    NUMBER_OF_CURRENT_REPORT_NUMERICAL_COLUMNS = 3

    REPORT_PERSIAN_NAME_TO_INCLUDE_IN_FIRST_SHEET = str()  # Sample Value:'تابستان 03'

    # These are some values that need to be saved right before working on the worksheets since if the numerical values
    # of sheets would be removed, they get missed
    LAST_SEASON_OTHER_SYSTEMS_TOTAL = str()
    LAST_SEASON_BMI_TOTAL = str()
    LAST_SEASON_MY_BMI_TOTAL = str()

    def calculate_all_coordinate_attributes(self):
        if self.__class__.counter_to_objects < 3:    # In monthly version of the automation => the first 2 sheets
            # should be ignored
            return None
        super().calculate_all_coordinate_attributes()   # The next worksheets can be processed with parent class
        # algorithm

    @classmethod
    def receive_report_persian_name_to_include_in_first_sheet_from_user(cls):
        """
        It Will Later Be Implemented In A GUI Style
        """

        cls.REPORT_PERSIAN_NAME_TO_INCLUDE_IN_FIRST_SHEET = input("\nEnter The Report Name In Persian To Include In "
                                                                  "First Sheet Of The Report(Sample Value:'بهار 03'): ")

    @classmethod
    def receive_header_names_of_numerical_columns_from_user(cls):
        """
        It Will Later Be Implemented In A GUI Style
        """

        print("\n##### Time Range Receive(Please Hit Enter Key After Each Month Name) #####")

        for i in range(cls.NUMBER_OF_LAST_REPORT_NUMERICAL_COLUMNS):
            item = input(f"Header Name '{i + 1}' = ")
            cls.LIST_HEADER_NAMES_OF_NUMERICAL_COLUMNS.append(item)

    @classmethod
    def extract_needed_values_from_first_sheet(cls, filename):
        """
        Important note for future: There is a possibility that the cell coordinates referenced in this method change as
        time passes, a dynamic algorithm to find them might be implemented later.

        ------------------------------------------------------------------------------------------------------------
        It is necessary to call this method before '.remove_values_of_cells_with_numerical_data()' is called. Why?
        the cells holding the values are formula-based. before having them all '0's, we have to save the statistical
        # data from last season.

        Question: Why the spreadsheet file is opened in this way in this function?
        Response: To have access to the calculated values of cells, we should set the 'data_only' flag to 'True'.

        :param filename: holds the filename of the spreadsheet
        """

        temp_workbook = load_workbook(filename, read_only=True, data_only=True)
        first_sheet = temp_workbook[temp_workbook.sheetnames[0]]    # I should work on the first worksheet only

        cls.LAST_SEASON_OTHER_SYSTEMS_TOTAL = first_sheet["B23"].value
        cls.LAST_SEASON_BMI_TOTAL = first_sheet["B24"].value
        cls.LAST_SEASON_MY_BMI_TOTAL = first_sheet["B25"].value

        temp_workbook.close()

        # test lines
        # print(f"cls.LAST_SEASON_OTHER_SYSTEMS_TOTAL = {cls.LAST_SEASON_OTHER_SYSTEMS_TOTAL}")
        # print(f"cls.LAST_SEASON_BMI_TOTAL = {cls.LAST_SEASON_BMI_TOTAL}")
        # print(f"cls.LAST_SEASON_MY_BMI_TOTAL = {cls.LAST_SEASON_MY_BMI_TOTAL}")

    def remove_values_of_cells_with_numerical_data(self):
        if self.__class__.counter_to_objects < 3:    # In monthly version of the automation => the first 2 sheets
            # should be ignored
            return None

        if self.sheet_has_issue is True:
            return None

        # clearing the numerical values of sequent time_based columns

        for row in self.list_of_numerical_cells_coordinates:
            for coordinate in row:
                self.sheet[coordinate].value = ""

    def delete_message_column_content(self):
        if self.__class__.counter_to_objects < 3:    # In monthly version of the automation => the first 2 sheets
            # should be ignored
            return None

        super().delete_message_column_content()

    def update_first_sheet_info(self):
        """
        The first worksheet of the report contains some parameters to be updated. The 'if-Block' has some simple(no
        algorithm-based) assignments that can be understood by taking a look at the report spreadsheet.
        """
        if self.__class__.counter_to_objects == 1:   # this method works only on the first worksheet

            # month names update
            self.sheet["B1"] = self.__class__.LIST_HEADER_NAMES_OF_NUMERICAL_COLUMNS[0]
            self.sheet["C1"] = self.__class__.LIST_HEADER_NAMES_OF_NUMERICAL_COLUMNS[1]
            self.sheet["D1"] = self.__class__.LIST_HEADER_NAMES_OF_NUMERICAL_COLUMNS[2]
            # ---------------------------------------

            # swap last report[quarter] name to the right cell and then insert this report[quarter] name
            self.sheet["C22"] = self.sheet["B22"].value
            self.sheet["B22"] = self.__class__.REPORT_PERSIAN_NAME_TO_INCLUDE_IN_FIRST_SHEET
            # ---------------------------------------

            # update values extracted in '.receive_filename_and_save_needed_values_on_first_sheet()' method
            self.sheet["C23"] = self.__class__.LAST_SEASON_OTHER_SYSTEMS_TOTAL
            self.sheet["C24"] = self.__class__.LAST_SEASON_BMI_TOTAL
            self.sheet["C25"] = self.__class__.LAST_SEASON_MY_BMI_TOTAL

    def header_date_fields_update(self):
        if self.__class__.counter_to_objects < 3:    # In monthly version of the automation => the first 2 sheets
            # should be ignored
            return None

        super().header_date_fields_update()

    @classmethod
    def determine_number_of_last_report_numerical_columns(cls):
        """
        :return: There is no no need for this method in this class.(The number always equals to 3)
        """
        return None

    @classmethod
    def determine_number_of_current_report_numerical_columns(cls):
        """
        :return: There is no no need for this method in this class.(The number always equals to 3)
        """
        return None

    def insert_or_delete_one_column_if_needed(self):
        """
        :return: There is no no need for this method in this class.
        """
        return None

    def update_font_and_style_of_the_new_numerical_column_if_inserted(self):
        """
        :return: There is no no need for this method in this class.
        """
        return None

    def update_all_formulas(self):
        if self.__class__.counter_to_objects < 3:    # In monthly version of the automation => the first 2 sheets
            # should be ignored
            return None

        super().update_all_formulas()
